import collections
import functools
import operator
import re
import warnings
from abc import ABC, abstractmethod
from idlelib.configdialog import changes

import pandas as pd
import numpy as np
from typing import Literal, Optional, Any
import itertools
from typing import Protocol, Type

from collections.abc import Callable
from enum import Enum, Flag

TABLE_DATA_MAP = {
    'fml': str, 'dependents': object, 'res_order': int, 
    'ftype': str, 'value': object, 'code': str
}

class XlFlags(Flag):
    ERROR_CLEAR = 1


class XlErrors(Enum):
  REF_ERROR = "#REF!"
  VALUE_ERROR = "#VALUE!"
  DIV_ZERO_ERROR = "#DIV/0!"
  NAME_ERROR = "#NAME?"
  NUM_ERROR = "#NUM!"
  NULL_ERROR = "#NULL!"
  GETTING_DATA_ERROR = "#GETTING_DATA"
  SPILL_ERROR = "#SPILL!"
  UNKNOWN_ERROR = "#UNKNOWN!"
  # Add more Excel error types here as needed

  def __str__(self):
    return self.value
  
  @property
  def code(self):
    id = list(self.__class__._value2member_map_.keys()).index(self.value)
    return f'Z{id}'


token_specification = [
    ('NUMBER', r'\d+(\.\d*)?'),  # Integer or decimal number
    ('ASSIGN', r'\='),  # Assignment operator
    ('SOP', r'^|&|<>'),  # Special operators
    ('OP', r'[+\-*/]'),  # Arithmetic operators
    ('COMMA', r','),  # Line endings
    ('ANCHOR', r'\:'),  # Line endings
    ('OPENP', r'\('),  # Line endings
    ('CLOSEP', r'\)'),  # Line endings
    ('SHEET', r"'[^']+'!"),  # Sheet names
    ('CELL', r'\$?[A-Z]\$?[1-9][0-9]*'),  # Identifiers
    ('FUNCTION', r'[A-Z]+'),  # Skip over spaces and tabs
    ('SKIP', r'[ %]+'),  # Skip over spaces and tabs
    ('MISMATCH', r'.'),  # Any other character
]
tokenizer = re.compile('|'.join('(?P<%s>%s)' % pair for pair in token_specification))

cell_pattern = re.compile(r"(?:'(?P<sht>.+?)'!)*(?P<col>\$?[A-Z]+)(?P<row>\$?[0-9]+)")
cell_address:Callable[[str], tuple[str, ...]] = lambda cell: tuple(x for x in cell_pattern.search(cell).groups()[::-1] if x)
rgn_pattern = re.compile(r"(?:'.+?'!)*\$?[A-Z]\$?[0-9]+(?::\$?[A-Z]\$?[0-9]+)?")
tbl_pattern = re.compile(r"(?:'(?P<sht>.+?)'!)*(?P<cell>.+)")
tbl_address = lambda tbl: tbl_pattern.match(tbl).groups()

ndx_sorter = lambda x: ('0000' + x.str.extract(r'(\d+)', expand=False)).str.slice(-4) + x.str.extract(r'([A-Z]+)', expand=False)

def excel_to_pandas_slice(excel_slice, axis, table_name, mask=None, withValues=False):
    sheet, lst_id = tbl_address(excel_slice)
    linf = f'{lst_id}:{lst_id}'.split(':', 1)[0]
    is_anchor = (linf[0] == '$' if axis == 0 else linf[1:].count('$') == 1)
    cell = row, col = cell_address(linf.replace('$', ''))
    mask = mask if not is_anchor else None
    if mask is None or (row not in mask and col not in mask):
        prefix = f"'{excel_slice}'" if sheet is None else f'"{excel_slice}"'
        pass
    elif row in mask or col in mask:
        infml = cell[col in mask]
        if sheet is None:
            prefix = ', '.join([f"'{excel_slice.replace(infml, key)}'" for key in mask])
        else:
            prefix = ', '.join([f'"\'{sheet}\'!{lst_id.replace(infml, key)}"' for key in mask])
    else:
        prefix = ''
    suffix = '.values' if withValues else ''
    py_term = f"{table_name}[{prefix}]{suffix}"
    return py_term

def pythonize_fml(fml: str, table_name: str, axis: None|Literal[0,1]=None, mask=None):
    '''
    Convierte una fórmula de Excel a una fórmula de Python
    :param fml: str. Fórmula de Excel
    :param table_name: str. Nombre de la tabla que contiene la fórmula
    :param axis: int. Eje de la tabla que se está procesando
    :return: str. Fórmula de Python
    '''

    pyfml = ''
    lst_id = ''
    fnc_stack = []
    for mo in re.finditer(tokenizer, fml):
        kind = mo.lastgroup
        token_chr = mo.group()
        nxt_char = fml[mo.end():].lstrip(' ')[0] if mo.end() < len(fml) else ''
        # print(f'{kind=}: {token_chr=}')
        match kind:
            case 'FUNCTION':
                fnc_name = token_chr if token_chr != 'IF' else 'WHERE'
                fnc_stack.append(fnc_name)
                pyfml += f'np.{fnc_name.lower()}'
            case 'ASSIGN':
                if pyfml.count('=') > 0 and pyfml[-1] not in '<>':
                    pyfml += '='
                pyfml += '='
            case 'OPENP':
                pyfml += '('
                fnc_stack.append('(')
            case 'CLOSEP':
                fnc_stack.pop()
                if fnc_stack and fnc_stack[-1] != '(':
                    fnc_name = fnc_stack.pop()
                    pyfml += f', axis={axis})' if fnc_name == 'SUM' else ')'
                else:
                    pyfml += ')'
            case 'ANCHOR':
                lst_id += token_chr
            case 'SHEET':
                lst_id = token_chr
            case 'CELL':
                prefix = token_chr.rstrip('0123456789')
                suffix = token_chr[len(prefix):]
                lst_id += f'{prefix}{suffix}'
                if ':' in lst_id or nxt_char != ':':
                    py_term = excel_to_pandas_slice(lst_id, axis, table_name, mask=mask, withValues='=' in pyfml)
                    pyfml += py_term
                    lst_id = ''
                    pass
            case 'NUMBER':
                if nxt_char == '%':
                    token_chr = f'({token_chr}/100.0)'
                pyfml += token_chr
            case 'SOP':
                if token_chr == '^':
                    token_chr = '**'
                elif token_chr == '&':
                    token_chr = '+'
                elif token_chr == '<>':
                    token_chr = '!='
                pyfml += token_chr
            case 'SKIP':
                pass
            case _:
                pyfml += token_chr
    return pyfml


def pass_anchors(coord1, coord2):
    anchors = ['$' if x[0] == '$' else '' for x in cell_pattern.match(coord1).groups()[1:]]
    sheet_name, *coords = cell_pattern.findall(coord2.replace('$', ''))[0]
    sheet = '' if not sheet_name else f"'{sheet_name}'!"
    return sheet + ''.join(x for tup in zip(anchors, coords) for x in tup)


def coords_from_range(rng):
    '''
    Encuentra el rango en que se revisaran las fórmulas.
    fml_col: str. Optional, si es None se supone que el rango de fórmulas es el que sigue al rng.
    rng: str. Rango de tabla que incluye concepto (ej: "B010:G227"
    '''
    rng_inf, rng_sup = rng.split(':')
    min_col, min_row = ord(rng_inf[0]) - ord('A') + 1, int(rng_inf[1:])
    max_col, max_row = (ord(rng_sup[0]) - ord('A') + 1) + 1, int(rng_sup[1:]) + 1
    return dict(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def data_in_range(openpyxl_sheet, ws_range:str):
    ws = openpyxl_sheet
    rgn_coords = coords_from_range(ws_range)
    rgn_fmls = {}
    rgn_values = {}
    rgn_map = [rgn_values, rgn_fmls]
    [
        operator.setitem(rgn_map[k], cell.coordinate, value.replace('=+', '=') if k == 1 else value)
        for row in ws.iter_rows(**rgn_coords)
        for cell in row
        if (value:=cell.value) and (k:=int(isinstance(value, str) and value[0] == '=')) < 2
    ]
    return rgn_fmls, rgn_values


class ExcelObject(ABC):
        _title: str
        id: Any
        parent: Optional['ExcelCollection']


        def __init__(self, object_name:str, *args, **kwargs):
            self._title = object_name
            pass

        @property
        def title(self) -> str:
            return self._title

        @title.setter
        def title(self, object_name:str):
            self._title = object_name
            pass


class ExcelCollection(ExcelObject, ABC):

    _title: str
    _objects: list[ExcelObject]
    object_class: Type[ExcelObject]
    parent: 'ExcelCollection'

    def __init__(self, parent, collection_name:str, obj_cls: Type[ExcelObject]):
        super().__init__(collection_name)
        self.parent = parent
        self._objects = []
        self.object_class = obj_cls
        pass

    @abstractmethod
    def next_id(self):
        pass

    def append(self, obj: ExcelObject):
        self._objects.append(obj)
        pass

    def index(self, excel_object: ExcelObject) -> int:
        tbl_name = excel_object.title
        return self.objectnames().index(tbl_name)

    def objectnames(self) -> list[str]:
        return [obj.title for obj in self._objects]

    def objects(self) -> list[ExcelObject]:
        return self._objects

    def __getitem__(self, object_name:str) -> ExcelObject:
        attr_name = 'id' if object_name.startswith('#') else 'title'
        object_name = object_name.strip('#')
        tbl = next(itertools.dropwhile(lambda x: getattr(x, attr_name) != object_name, self._objects))
        assert getattr(tbl, attr_name) == object_name, f'{object_name} not found'
        return tbl

    def move_object(self, object_name:str, ins_point:int = None):
        ndx = self.objectnames().index(object_name)
        excel_object = self._objects.pop(ndx)
        ins_point = ins_point or len(self._objects)
        self._objects.insert(ins_point, excel_object)
        pass

    def _create_object(self, object_name:str|None, ins_point:int|None = None) -> ExcelObject:
        obj = self.object_class(self, object_name)
        if ins_point:
            self.move_object(obj.title, ins_point)
        return obj


class ExcelWorkbook(ExcelCollection):

    def __init__(self, workbook_name: str):
        super().__init__(None, workbook_name, ExcelWorksheet)
        self.links = collections.defaultdict(set)
        self.parameters = collections.defaultdict(set)
        self.count = -1
        pass

    def next_id(self):
        self.count += 1
        return f'sheet{"" if self.count == 0 else self.count}'

    # def create_worksheet(self, sheet_name:str, ins_point:int = None) -> "ExcelWorksheet":
    #     return self._create_object(sheet_name, ins_point)

    create_worksheet = ExcelCollection._create_object

    sheetnames = property(ExcelCollection.objectnames)

    sheets = property(ExcelCollection.objects)


class ExcelWorksheet(ExcelCollection):

    def __init__(self, parent:ExcelWorkbook, sheet_name:str|None):
        id = parent.next_id()
        sheet_name = sheet_name or id
        super().__init__(parent, sheet_name, ExcelTable)
        self.parent.append(self)
        self.id = id
        self.count = -1
        self._param_map = {}
        self._data_rgn = None
        pass


    def propagate_error(self, xl_error: XlErrors, codes: list[str] | None = None, reg_value=None):
        assert reg_value is None or not isinstance(reg_value, XlErrors), 'reg_value must be None or not a XlErrors'
        def register_code(tbl, xl_error, codes):
                df = tbl.data
                err_cell, err_code, err_value = next(zip(*self.register_links(tbl, [f"'{self.title}'!{xl_error.code}"], default_value=xl_error)))
                try:
                    error_rec = df.loc[err_cell].copy()
                except:
                    # Si no existe código del error en la tabla, se crea un nuevo registro para el error
                    error_rec = pd.Series(
                        dict(fml=None, dependents=set(), res_order=0, ftype='$', value=err_value, code=err_code), 
                        dtype=object
                    )
                # Se agrega el código de la celda al campo dependents del registro del error
                error_rec.dependents.update(tbl_address(x)[-1] for x in codes)
                df.loc[err_cell] = error_rec
        
        def unregister_code(tbl, xl_error, codes):
                df = tbl.data
                err_cell, err_code, err_value = next(zip(*self.register_links(tbl, [f"'{self.title}'!{xl_error.code}"], default_value=xl_error)))
                error_rec = df.loc[err_cell].copy()
                # Se agrega el código de la celda al campo dependents del registro del error
                error_rec.dependents.difference_update(tbl_address(x)[-1] for x in codes)
                df.loc[err_cell] = error_rec
                if not error_rec.dependents:
                    df.drop(err_cell)
                    wb.links[err_code].difference_update([f"'{tbl.parent.id}'!{tbl.id}"])
                    if not wb.links[err_code]:
                        wb.links.pop(err_code)
                        tbl.parent._param_map.pop(err_cell)


        wb = self.parent
        to_process = {}
        if codes is None:
            # Se quiere registrar todos los códigos de las celdas que tienen un error
            # por el momento esto se hace solo en REF_ERROR ya que los otros salen del cálculo
            # de las celdas.
            for tbl in self.tables:
                df = tbl.data
                try:
                    dependents = df.loc[xl_error.code, 'dependents']
                except KeyError:
                    continue
                to_process.setdefault(tbl, []).extend(f"'{tbl.parent.id}'!{x}" for x in dependents)
        else:
            # Si codes no es None, se debe registrar el codigo de la celda en los dependents de la
            # tabla correspondiente
            
            
            for code in codes:
                ws_id, tbl_id, _ = cell_pattern.match(code).groups()
                tbl = wb['#' + ws_id]['#' + tbl_id]
                # Se consigna para procesamiento en la clave de la tabla correspondiente
                to_process.setdefault(tbl, []).append(code)
            fnc = register_code if reg_value is None else unregister_code
            [fnc(tbl, xl_error, codes) for tbl, codes in to_process.items()]

        propagate_error = {}
        while to_process:
            # print(f'{to_process=}')
            tbls = list(to_process.keys())
            while tbls:
                tbl = tbls.pop()
                codes = to_process.pop(tbl)
                dependents = [tbl.encoder('decode', tbl_address(x)[1]) for x in codes]
                propagate_error.setdefault(tbl, []).extend(dependents)
                # print(f'ws = {tbl.parent.title}, tbl = {tbl.title}')
                # print(sorted(dependents, key=lambda x: '{0: >4s}{1: >2s}'.format(*cell_address(x))))
                all_dep = set(tbl.all_dependents(dependents).dependent)
                # print(f'{all_dep=}')
                for dep in all_dep:
                    sht_id, tbl_id = cell_pattern.match(dep).groups()[:2]
                    tbl = wb['#' + sht_id]['#' + tbl_id]
                    to_process.setdefault(tbl, []).append(dep)
        reg_value = reg_value or xl_error
        for tbl, cells in propagate_error.items():
            tbl.data.loc[cells, 'value'] = reg_value

    def insert(self, cell_slice, from_delete=False):
        excel_slice = f'{cell_slice}:{cell_slice}'.split(':', 2)[:2]
        slice_type = cell_slice.replace(':', '')
        is_numeric = slice_type.isnumeric()
        is_alpha = slice_type.isalpha()
        is_cell = is_numeric and is_alpha
        if is_cell:    # insert columns / rows
            return
        if is_numeric:    # insert rows
            field = 'row_offset'
            fnc = int
        else:                           # insert columns
            field = 'col_offset'
            fnc = ord

        nitems = fnc(excel_slice[1]) - fnc(excel_slice[0]) + 1
        kwargs = {field: (-nitems if from_delete else nitems)}
        kwargs['disc_cell'] = f'A{excel_slice[0]}' if is_numeric else f'{excel_slice[0]}1'

        ws = self
        wb = ws.parent

        for tbl in self.tables:
            if (data_rng := tbl.offset_rng(tbl.data_rng, **kwargs)) == tbl.data_rng:
                continue

            df = tbl.data

            # modificación data range
            cells = tbl.cells_in_data_rng(df.index.tolist())
            tbl.data_rng = data_rng

            # modificación del índice del tbl.data frame (cells labels)
            cells_map = tbl.offset_rng(cells, **kwargs)
            tbl.data.rename(index=cells_map, inplace=True)

            # modificación enlaces externos a cells modificadas
            mask = df.index.isin(cells_map.values())
            external_links = wb.links.keys() & set(f"'{ws.id}'!{x}" for x in df.loc[mask].code)
            codes = [tbl_address(code)[-1] for code in external_links]
            codes_map = df.loc[df.code.isin(codes)].code.to_dict()

            to_broadcast = {
                f"'{ws.id}'!{code}": f"'{ws.title}'!{cell}" 
                for cell, code in codes_map.items()
            }
            ws.broadcast_changes(to_broadcast, field='cell')
        self._data_rgn = None

    def delete(self, cell_slice):
        excel_slice = f'{cell_slice}:{cell_slice}'.split(':', 2)[:2]
        slice_type = cell_slice.replace(':', '')
        is_numeric = slice_type.isnumeric()
        is_alpha = slice_type.isalpha()
        is_cell = is_numeric and is_alpha
        (rmin, cmin),  (rmax, cmax) = map(cell_address, self.data_rng.split(':'))
        if is_cell:    # insert columns / rows
            return
        if is_numeric:    # insert rows
            field = 'row_offset'
            fnc = int
            cmin, cmax = map(ord, (cmin, cmax))
            rmin, rmax = map(int, excel_slice)
        else:                           # insert columns
            field = 'col_offset'
            fnc = ord
            cmin, cmax = map(ord, excel_slice)
            rmin, rmax = map(int, (rmin, rmax))

        nitems = fnc(excel_slice[1]) - fnc(excel_slice[0]) + 1
        kwargs = {field: nitems}
        kwargs['disc_cell'] = f'A{excel_slice[0]}' if is_numeric else f'{excel_slice[0]}1'

        all_cells = [f'{chr(col)}{row}' for col in range(cmin, cmax + 1) for row in range(rmin, rmax + 1)]
        ws = self
        wb = ws.parent

        for tbl in self.tables:
            if not (cells := tbl.cells_in_data_rng(all_cells)):
                continue
            df = tbl.data
            cells = list(set(cells) & set(df.index))
            codes = df.loc[cells].code.tolist()
            off_cells = tbl.offset_rng(cells, **kwargs).values()
            off_codes = df.loc[off_cells].code.tolist()
            rgn_codes = dict(zip(codes, off_codes))
            err_cell, err_code, err_value = next(zip(*self.register_links(tbl, [f"'{ws.title}'!{XlErrors.REF_ERROR.code}"], default_value=XlErrors.REF_ERROR)))
            changes = {code: err_code for code in codes}
            tbl.set_field(changes, field='code', rng_codes=rgn_codes)
            changes = {f"'{ws.id}'!{code}": err_code for code in codes}
            self.broadcast_changes(changes, field='code')

        assert not df.index.isin(cells).any(), 'Insert: Not all cells has been displaced'
        self.insert(cell_slice, from_delete=True)
        self.propagate_error(XlErrors.REF_ERROR)
        for tbl in self.tables:
            if not (cells := tbl.cells_in_data_rng(all_cells)):
                continue
            cells = list(set(cells) & set(tbl.data.index))
            df = tbl.data.loc[cells]
            codes = df.code.tolist()
            mask =  ~df.dependents.isna()
            all_dependents = set(itertools.chain(*df.loc[mask].dependents.tolist()))
            all_dependents = all_dependents - set(codes)
            mask = tbl.data.code.isin(all_dependents) & ~tbl.data.value.isin(list(XlErrors))
            if mask.any():
                changed = tbl.data.loc[mask].code.tolist()
                tbl.changed.extend(changed)
                tbl.recalculate(recalc=True)
        pass


    @property
    def data_rng(self):
        if self._data_rgn is None:
            cells = list(map(cell_address, [x for x in self._param_map.keys() if not x.startswith('Z')]))
            for tbl in self.tables:
                cells.extend(map(cell_address, tbl.data_rng.split(':')))
            rows, cols = zip(*cells)
            rows = [int(x) for x in rows]
            cols = [f'{x: >2s}' for x in cols]
            r_min, r_max = min(rows), max(rows)
            c_min, c_max = map(str.strip, (min(cols), max(cols)))
            self._data_rng = f'{c_min}{r_min}:{c_max}{r_max}'
        return self._data_rng
    
    def _repr_html_(self):
        (rmin, cmin),  (rmax, cmax) = map(cell_address, self.data_rng.split(':'))
        rmin, rmax = int(rmin), int(rmax)
        cmin, cmax = ord(cmin), ord(cmax)
        all_cells = [f'{chr(col)}{row}' for row in range(rmin, rmax + 1) for col in range(cmin, cmax + 1)]
        t = pd.Series('', index=all_cells)
        if (params := [x for x in self.parameters() if not x.startswith('Z')]):
            t[params] = self.parameters(*params)
        for tbl in self.tables:
            cells_rng = tbl._cell_rgn(tbl.data_rng)
            mask = tbl.data.index.isin(cells_rng)
            t[tbl.data.index[mask]] = tbl.data.value[mask]
        data = ExcelTable.excel_table(t)
        return data._repr_html_()


    def next_id(self):
        self.count += 1
        return chr(ord('A') + self.count)

    create_table = ExcelCollection._create_object

    @property
    def tablenames(self):
        return self.objectnames()

    @property
    def tables(self):
        return self._objects

    def remove(self, tbl:'ExcelTable'):
        del tbl

    def register_tbl(self, tbl: 'ExcelTable'):
        cells_to_link = self.parameters()
        if cells_to_link and (links_in_tbl := tbl.cells_in_data_rng(list(cells_to_link))):
            old_codes = [self.parameter_code(cell) for cell in links_in_tbl]
            code_links = [self.parameter_code(cell) for cell in tbl.data.loc[links_in_tbl, 'code']]
            changes = dict(zip(old_codes, code_links))
            self.broadcast_changes(changes, field='cell')
            links = self.parent.links
            for old_code, code in changes.items():
                links[code] = links.pop(old_code)
            [self._param_map.pop(key) for key in links_in_tbl]
        pass

    def register_links(self, xtable: 'ExcelTable', to_link: list[str], default_value=0) -> tuple[list[str], list[str], list[Any]]:
        to_link = ['!'.join(f"'{self.title}'!{x}".split('!')[-2:]) for x in to_link]
        to_link = sorted(to_link, key=lambda x: '{2: >40s}!{0: >4s}{1}'.format(*cell_address(x)))

        links = self.parent.links
        answ = []
        for cell in to_link:
            sheet_name, cell_coord = tbl_address(cell)
            sheet: ExcelWorksheet = self.parent[sheet_name]
            tbl =sheet.associated_table(cell)
            try:
                code, value = tbl.data.loc[cell_coord, ['code', 'value']]
            except AttributeError:
                pvalue = default_value
                value = sheet._param_map.setdefault(cell_coord, pvalue)
                code = 'Z' + cell_coord
            code = f"'{sheet.id}'!{code}"
            cell = f"'{sheet.title}'!{cell_coord}" if sheet.id != self.id else cell_coord
            answ.append((cell, code, value))
            links[code].add(f"'{self.id}'!{xtable.id}")

        cells, code_links, values = map(list, zip(*answ))
        return cells, code_links, values
    
    def parameter_code(self, param_name:str) -> str:
        return f"'{self.id}'!Z{param_name}"

    def parameters(self, *args, **kwargs):
        '''
        Parameters are the links in the existing tables to this sheet cells that do not belong
        to any defined table.
        '''
        # If args are provided, return the corresponding values for the parameters
        if args:
            return [self._param_map.get(arg) for arg in args]
        # If keyword arguments are provided, update the values of the parameter and broadcast changes
        elif kwargs:
            keys = kwargs.keys() & self._param_map.keys()
            self._param_map.update((key, kwargs[key]) for key in keys)
            values = {self.parameter_code(key): kwargs[key] for key in keys}
            self.broadcast_changes(values, field='parameter')
        # If no arguments are provided, return the list of parameters of the sheet.
        else:
            return list(self._param_map.keys())

    def cell_values(self, input_cells: list[str]):
        value_items = []
        s = pd.Series(input_cells)
        sheet_groups = (
            s
            .where(s.str.contains('!'), f"'{self.title}'!" + s)
            .str.extract(tbl_pattern, expand=True)
            .set_index('cell')
            .groupby(by='sht')
            .groups
            )
        
        for sht, items in sheet_groups.items():
            n = 0
            ws = self.parent[sht]
            ws_name = ws.title
            links = items.tolist()
            while links and n < len(ws.tables):
                tbl: ExcelTable = ws.tables[n]
                if cells:=tbl.cells_in_data_rng(links):
                    value_items.extend(
                        (f"'{ws_name}'!{cell}", value)
                        for cell, value in
                        tbl.data.loc[cells, 'value'].to_dict().items()
                    )
                    links = list(set(links) - set(cells))
                n += 1
            if (parameters := set(links) & set(ws.parameters())):
                pvalues = ws.parameters(*parameters)
                value_items.extend([(f"'{ws_name}'!{cell}", value) for cell, value in zip(parameters, pvalues)])
            if (links := set(links) - set(parameters)):
                value_items.extend([(f"'{ws_name}'!{cell}", None) for cell in links])
        value_items = sorted(value_items, key=lambda x: '{2: >40s}!{0: >4s}{1}'.format(*cell_address(x[0])))
        return map(list, zip(*value_items))

    def broadcast_changes(self, changes: dict[str, Any], *, field: Literal['value', 'parameter', 'cell'] = 'value'):
        ws_id = self.id
        links = self.parent.links
        keys = list(changes.keys() & links.keys())
        if keys:
            tbl_df = (
                pd.DataFrame(
                    [{tbl: 1 for tbl in links[key]} for key in keys],
                    index=keys
                )
                .fillna(0)
            )
            columns = [key for key in sorted(tbl_df.columns, key=lambda x: tbl_address(x))]
            for tbl_addr in columns:
                if not any(mask:=tbl_df[tbl_addr] == 1):
                    continue
                cells = tbl_df[mask].index.tolist()
                ws_id, tbl_id = map(lambda x: f'#{x}', tbl_address(tbl_addr))
                ws = self.parent[ws_id]
                tbl = ws[tbl_id]
                match field:
                    case 'value' | 'parameter':
                        cells = tbl.encoder('decode', pd.Series(cells, index=cells)).to_dict()
                        mapper = {key: changes[cell] for cell, key in cells.items()}
                        tbl.set_values(mapper, field=field, recalc=True)
                    case 'cell':  # field == 'cell'
                        mapper = {cell: changes[cell].replace(f"'{ws.title}'!", '') for cell in cells}
                        tbl.set_field(mapper, field='cell')
                    case 'code':  # field == 'code'
                        mapper = {cell: changes[cell] for cell in cells}
                        tbl.set_field(mapper, field='code')
                    case _:
                        pass
        pass

    def associated_table(self, cell:str, scope:Literal['cell', 'parametr']='cell') -> ExcelObject| list[ExcelObject] | None:
        match scope:
            case 'cell':
                tables: list[ExcelObject] = self._objects
                for tbl in tables:
                    if cell in tbl:
                        return tbl
            case 'parameter':
                wb = self.parent
                code = self.parameter_code(cell)
                tbl_codes = self.parent.links[code]
                tbl_pairs = [tbl_address(tbl_code) for tbl_code in tbl_codes]
                tbls = [wb[f'#{sht_id}'][f'#{tbl_id}'] for sht_id, tbl_id in tbl_pairs]
                return tbls
        return None


class ExcelTable(ExcelObject):

    def __init__(self, parent:ExcelWorksheet, tbl_name:str|None, table_rng:str, fmls: dict[str,str]|None, values:dict[str, Any]|None, recalc:bool=False):

        super().__init__(tbl_name)
        self.parent: ExcelWorksheet = parent
        self.id = parent.next_id()
        self.count = 0
        self.data = (
            pd.DataFrame(columns=TABLE_DATA_MAP.keys())
            .astype(TABLE_DATA_MAP)
        )
        self.data_rng = table_rng
        self.needUpdate: bool = False
        self.changed = []
        parent.append(self)
        if fmls:
            self.set_fmls(fmls, values, recalc)
        self.set_values(values, recalc=recalc)
        pass

    def next_id(self):
        self.count += 1
        return f'{self.id}{self.count}'

    def normalize_data(self, df):
        df = (
            df
            .assign(code=lambda db: [self.next_id() for n in range(1, len(db.index) + 1)])
            .assign(fml= lambda db: self.encoder('encode', db['fml'], df=db))
            .assign(dependents=lambda db: db.dependents.apply(lambda x: {db.code[term] for term in x} if x is not np.nan else x))
        )
        return df

    def set_fmls(self, fmls: dict[str,str]|None, values:dict[str, Any]|None, recalc:bool=False):

        df = (
            pd.DataFrame.from_dict(fmls, orient='index', columns=['fml'], dtype=TABLE_DATA_MAP['fml'])
            .pipe(self.fml_dependents)
            .pipe(self.res_order)
            .drop(columns='independents')
            .pipe(self.formula_type)
            .assign(value=0)
            .pipe(self.normalize_data)
        )
        self.data = (
            pd.concat([self.data, df])
            .sort_index()
        )
        self.data.index.rename('cell', inplace=True)
        df = self.data
        df = self.data
        self.parent.register_tbl(self)
        values = values or {}
        if links:=self.links():
            ws = self.parent
            links, code_links, link_values = ws.register_links(self, links)
            # Asignación de code para los enlaces (links)
            old_codes = df.loc[links, 'code'].tolist()
            changes = dict(zip(old_codes, code_links))
            self.set_field(changes, field='code')
            # Asignación de valores para los enlaces (links)
            df.loc[links, 'value'] = link_values
            self.changed.extend(code_links)

    def cells_in_data_rng(self, cells: list[str]) -> list[str]:
        (rmin, cmin),  (rmax, cmax) = map( cell_address, self.data_rng.split(':'))
        ws_name = self.parent.title
        s = pd.Series(cells)
        s = s.where(s.str.contains('!'), f"'{ws_name}'!" + s)
        db = s.str.extract(cell_pattern, expand=True)
        mask = (db.sht == f'{ws_name}') & db.row.astype(int).between(int(rmin), int(rmax)) & db.col.between(cmin, cmax)
        return pd.Series(cells, index=mask).loc[True].tolist() if mask.any() else []

    def links(self):
        cells = self.data.index.tolist()
        cells_in_rgn = self.cells_in_data_rng(cells)
        cell_links = list(set(cells) - set(cells_in_rgn))
        return cell_links

    def __del__(self):
        ndx = self.parent.index(self)
        self.parent._objects.pop(ndx)

    def set_field(self, changes: dict[str, Any], *, field: Literal['code', 'cell']= 'code', rng_codes: dict[str, str] | None = None):
        assert (df := self.data) is not None, 'Table not initialized'
        rng_codes = rng_codes or {}
        ws = self.parent
        if field == 'code':
            old_codes = list(changes.keys())
            if not (gmask := df.code.isin(old_codes)).any():
                return
            old_codes, dependents = zip(*df.loc[gmask, ['code', 'dependents']].values)
            code_links = [changes[key] for key in old_codes]
            is_delete_operation = len(set(code_links)) == 1 and code_links[0].endswith(XlErrors.REF_ERROR.code)
            for dependents_set, old_code, code in zip(dependents, old_codes, code_links):
                if not isinstance(dependents_set, set):
                    continue
                mask = df.code.isin(dependents_set)
                if not is_delete_operation:
                    pattern = r"(?:'({0})'!)*(\$?{1})(\$?{2})".format(*cell_pattern.match(old_code).groups())
                    replacement = lambda m: pass_anchors(m[0], code)
                else:
                    pattern = rgn_pattern
                    def replacement(m):
                        pattern = r"(?:'({0})'!)*(\$?{1})(\$?{2})".format(*cell_pattern.match(old_code).groups())
                        if ':' in m[0]:
                            if all(x in rng_codes for x in m[0].split(':')):
                                return code
                            return re.sub(pattern, lambda m: pass_anchors(m[0], rng_codes[old_code]), m[0])
                        elif re.match(pattern, m[0]):
                            return code
                        else:
                            return m[0]

                df.loc[mask, 'fml'] = df[mask].fml.str.replace(
                    pattern,
                    replacement,
                    regex=True
                )
            df.loc[gmask, 'code'] = code_links
            if is_delete_operation and df.code.tolist().count(code_links[0]) > 1:
                err_code = code_links[0]
                err_cell = err_code.replace(f"'{ws.id}'!Z", '')
                # Se eliminan las filas duplicadas que resultan de la eliminación de enlaces a 
                # celdas inexistentes
                df.drop_duplicates(subset=['code'], inplace=True, keep=False)
                # Se identifican los errors origens
                mask = df.fml.fillna('').str.contains(err_code)
                error_origens = set(df.loc[mask, 'code'])
                df.loc[err_cell, ['code', 'fml','dependents', 'res_order', 'ftype', 'value']] = [err_code, None, error_origens, 0, '$', XlErrors.REF_ERROR]
                # Se eliminan en el campo dependents las referencias a los códigos a ser eliminados
                if (mask := df.dependents.apply(lambda x: bool((x if isinstance(x, set) else set()) & set(old_codes)))).any():
                    df.loc[mask, 'dependents'] = df.loc[mask].dependents.apply(lambda x: set(x) - set(old_codes))

        else:  # field == 'cell'
            mask = df.code.isin(changes.keys())
            keys = df.loc[mask, 'code'].apply(changes.get)
            df.rename(index=keys, inplace=True)
        pass

    def set_values(self, values: dict[str, Any], field: Literal['value', 'parameter']='value', recalc:bool=False):
        assert (df := self.data) is not None, 'Table not initialized'
        keys = self.cells_in_data_rng(values.keys())
        if not keys:
            return
        if (to_init := list(set(values.keys()) - set(df.index))):
            value_rec = pd.Series(
                dict(fml=None, dependents=None, res_order=0, ftype='$', value=None, code=''), 
                dtype=object
            )
            df = pd.concat(
                [
                    df,
                    pd.DataFrame(value_rec.to_dict(), index=pd.Index(to_init, name='cell'))
                ]
            )
            df.loc[to_init, 'code'] = [f'{self.next_id()}' for _ in range(len(to_init))]
            self.data = df
        values = {key: values[key] for key in keys}
        changed = keys
        if (parameters:= values.keys() & set(self.parent.parameters()) and field == 'values'):
            warnings.warn(
                f"You are modifying this parameter(s): {parameters}. "
                f"At this level it only modifies the table in which they are set."
                f"To be valid across the workbook, try at the worksheet level with the function "
                f"ws.parameters(param1=value1, param2=value2)",
                UserWarning
            )
        if changed or self.changed:
            self.changed.extend(df.loc[changed].code.tolist())
            df.loc[changed, 'value'] = pd.Series(values, dtype=TABLE_DATA_MAP['value'])
        self.recalculate(recalc)
        pass

    def get_cells_to_calc(self, changed):
        to_process = set()
        for res_order, changed in self.cells_to_calc(changed):
            to_process.update(changed)
        return sorted(to_process, key=lambda x: '{0: >4s}{1}'.format(*cell_address(x)))

    def ordered_formulas(self, order, feval=False):
        assert self.data is not None, 'Table not initialized'
        gmask = (self.data.code.isin(order) & (self.data.res_order > 0)).tolist()
        df = (
            self.data
            .reset_index()
            .set_index('code')
        )
        formulas = []
        for (res_order, ftype), cells in df[gmask].groupby(['res_order', 'ftype']):
            cells.fml = self.encoder('decode', cells.fml, df=df)
            if ftype == '$' or len(cells) == 1:
                formulas.extend(
                    [
                        (*cell_address(rec.cell), rec.fml)
                        for code in cells.index.tolist()
                        if (rec := cells.loc[code]).any()
                    ]
                )
            else:
                mask = cells.cell.tolist()
                row, col = map(set, zip(*[cell_address(cell) for cell in mask]))
                ftype, cell_range = (row.pop(), sorted(col)) if len(row) == 1 else (col.pop(), sorted(row))
                formulas.append(
                    (ftype, cell_range, cells.iloc[0].fml)
                )
            pass

        pyfmls = []
        for frst_item, scnd_item, fml in formulas:
            pyfml = self.formula_translation(frst_item, scnd_item, fml, feval=feval)
            pyfmls.extend(pyfml)
            # print(f'{frst_item} {scnd_item} {fml} ==> {pyfml}')

        return pyfmls

    def formula_translation(self, frst_item, scnd_item, fml, table_name='tbl', feval=False):
        '''
        Traduce una fórmula de Excel a una fórmula de Python
        :param fml: str. Fórmula de Excel
        :param table_name: str. Nombre de la tabla que contiene la fórmula
        :return: str. Fórmula de Python
        '''

        match (frst_item, scnd_item, fml):
            case (str(frst_item), str(scnd_item), fml):  # cell fml
                # to_pythonize = f'{scnd_item}{frst_item}{fml}'
                mask = [(f'{scnd_item}{frst_item}', scnd_item)]
                # to_pythonize = f'{scnd_item}{frst_item}{fml}'
                mask = [(f'{scnd_item}{frst_item}', scnd_item)]
                axis = None            # To avoid FutureWarning
            case (str(frst_item), list(scnd_item), fml) if frst_item.isnumeric():  # row fml
                # to_pythonize = f'{scnd_item[0]}{frst_item}{fml}'
                scnd_item = sorted(scnd_item)
                mask = [(f'{col}{frst_item}', col) for col in scnd_item]
                # to_pythonize = f'{scnd_item[0]}{frst_item}{fml}'
                scnd_item = sorted(scnd_item)
                mask = [(f'{col}{frst_item}', col) for col in scnd_item]
                axis = 0
            case (str(frst_item), list(scnd_item), fml) if frst_item.isalpha():  # column fml
                # to_pythonize = f'{frst_item}{scnd_item[0]}{fml}'
                scnd_item = sorted(scnd_item, key=lambda x: int(x))
                mask = [(f'{frst_item}{row}', row) for row in scnd_item]
                # to_pythonize = f'{frst_item}{scnd_item[0]}{fml}'
                scnd_item = sorted(scnd_item, key=lambda x: int(x))
                mask = [(f'{frst_item}{row}', row) for row in scnd_item]
                axis = 1
            case _:  # (list(frst_item), list(scnd_item), fml)    # range fml
                to_pythonize = ''
                mask = None
                axis = None
        py_fml = pythonize_fml(fml, table_name=table_name, axis=axis)
        frst_item = mask[0][1]
        if feval:
            py_fml = py_fml.lstrip('=')
            return [(cell, py_fml.replace(frst_item, x)) for cell, x in mask]
        return [cell + py_fml.replace(frst_item, x) for cell, x in mask]

    def formula_type(self, df):
        order = (
            df
            .loc[df.res_order > 0, ['fml', 'res_order']]
            .reset_index(names='cell')
            .set_index('res_order')
            .sort_index()
        )

        def fml_equiv(m, item, frst_item):
            sheet, col, row = m.groups()
            if (frst_item.isnumeric() and row[0] == '$') or (frst_item.isalpha() and col[0] == '$'):
                return m[0]
            return m[0].replace(item, frst_item)

        rgn_fmls = lambda col, row: df.loc[f'{col}{row}' if col.isalpha() else f'{row}{col}', 'fml']

        formulas = []

        for res_order in order.index.unique():
            batch = order.loc[order.index == res_order, ['cell', 'fml']].values.tolist()
            batch = sorted(batch, key=lambda x: '{0: >4s}{1}'.format(*cell_address(x[0])))

            maps = [collections.defaultdict(list), collections.defaultdict(list), collections.defaultdict(list)]
            rmap, next = maps[:2]
            [
                rmap[row].append(col)
                for cell_coord, fml in batch
                if (cell_tple := cell_address(cell_coord)) and (row := cell_tple[0]) and (col := cell_tple[1])
            ]
            # Rows with only one associated col are transferred to the cols map.
            keys = list(rmap.keys())
            [next[col].append(row) for row in keys if len(rmap[row]) == 1 and (col := rmap.pop(row)[0])]

            for _ in range(2):
                for row in sorted(rmap.keys(), key=lambda x: len(rmap[x])):
                    columns = rmap.pop(row)
                    while columns:
                        frst_col, *columns = sorted(columns)
                        test_fml = rgn_fmls(frst_col, row)
                        mask = [
                            col
                            for col in columns
                            if test_fml == cell_pattern.sub(lambda m: fml_equiv(m, item=col, frst_item=frst_col),
                                                            rgn_fmls(col, row))
                        ]
                        if not mask:
                            next[frst_col].append(row)
                        else:
                            mask = [frst_col] + mask
                            formulas.extend([(f'{col}{row}' if col.isalpha() else f'{row}{col}', row) for col in mask])
                            columns = [col for col in columns if col not in mask]

                rmap, next = maps[1:]
                # cols with only one associated row are transferred to the single_list
                keys = list(rmap.keys())
                [
                    next[row].append(col)
                    for col in keys
                    if len(rmap[col]) == 1 and (row := rmap.pop(col)[0])
                ]

            formulas.extend([(f'{col}{row}', '$') for row, mask in next.items() for col in mask])
        ftype = pd.Series(dict(formulas), dtype=TABLE_DATA_MAP['ftype'])
        df.loc[ftype.index, 'ftype'] = ftype
        return df

    def cells_to_calc(self, init_changed):
        df = (
            self.data
            .reset_index()
            .set_index('code')
        )
        to_report = df.loc[init_changed, ['res_order']].groupby(by='res_order').groups
        while True:
            try:
                k_min = min(to_report.keys())
            except ValueError:
                break
            changed = list(to_report.pop(k_min))
            changed = (yield (k_min, changed)) or changed
            if changed:
                dependents = df.loc[changed].dependents
                mask = ~dependents.isna()
                if mask.any():
                    changed = list(functools.reduce(lambda t, e: t.union(e), dependents[mask], set()))
                    grouped_changed = df.loc[changed, ['res_order']].groupby(by='res_order').groups.items()
                    [to_report.setdefault(k, set()).update(v) for k, v in grouped_changed]

    def evaluate(tbl, formulas):
        answer = []
        for cell, py_fml in formulas:
            try:
                value = eval(py_fml, globals(), locals())
            except ZeroDivisionError as e:
                value = XlErrors.DIV_ZERO_ERROR
            except TypeError as e:
                value = XlErrors.VALUE_ERROR
            except NameError as e:
                value = XlErrors.NAME_ERROR
            except Exception as e:
                value = XlErrors.UNKNOWN_ERROR

            try:
                value = value.flatten()[0]
            except AttributeError:
                pass
            answer.append((cell, value))
        return answer

    def recalculate(tbl, recalc:bool=False):
        if recalc:
            ws = tbl.parent
            df = tbl.data
            # Se obtienen los códigos de las celdas que tienen un error
            error_origens = set()
            if (mask := df.index.isin([x.code for x in list(XlErrors)])).any():
                [error_origens.update(x) for x in df[mask].dependents.to_list()]
            changed, tbl.changed = tbl.changed, []
            values = {}
            changed_cells, f_changed = tbl.cells_to_calc(changed), None
            while True:
                try:
                    res_order, changed = changed_cells.send(f_changed)
                except StopIteration:
                    break
                mask = tbl.data.code.isin(changed)
                s0 = pd.Series({
                    code: value 
                    for code, value in tbl.data.loc[mask, ['code', 'value']].values
                    if code in error_origens or not isinstance(value, XlErrors)
                }, dtype=TABLE_DATA_MAP['value'])
                changed = f_changed = s0.index.tolist()
                if not f_changed:
                    continue
                if not res_order:
                    values.update(s0.to_dict())
                    continue
                formulas = tbl.ordered_formulas(changed, feval=True)
                items = tbl.evaluate(formulas)
                cells, vals = map(list, zip(*items))
                s1 = pd.Series(vals, index=tbl.data.loc[cells].code, dtype=TABLE_DATA_MAP['value'])     # New values
                errors_mask = s1.isin(list(XlErrors))
                mask = ~errors_mask & (s0 != s1[s0.index])
                # Celdas que al recalcular se generan nuevos errores
                if errors_mask.any():
                    errors = s1[errors_mask]
                    to_propagate = {}
                    [
                        to_propagate.setdefault(y, []).append(f"'{tbl.parent.id}'!{x}")
                        for x, y in errors.items()
                    ]
                    for err, codes in to_propagate.items():
                        tbl.parent.propagate_error(err, codes=codes)
                    pass
                vals = s1[mask]
                f_changed = vals.index.tolist()
                if not f_changed:
                    continue
                # Celdas con error que al recalcular se elimina el error
                if (error_clear := vals[vals.index.isin(error_origens)].index.tolist()):
                    errors = s0[error_clear].to_dict()
                    xl_errors = {}
                    [xl_errors.setdefault(y, []).append(f"'{tbl.parent.id}'!{x}") for x, y in errors.items()]
                    for err, codes in xl_errors.items():
                        tbl.parent.propagate_error(err, codes=codes, reg_value=XlFlags.ERROR_CLEAR)
                values.update(vals.to_dict())
                vals.rename(index=lambda x: tbl.encoder('decode', x), inplace=True)
                tbl.data.loc[vals.index, 'value'] = vals

            ws = tbl.parent
            values = dict((f"'{ws.id}'!{key}", value) for key, value in values.items())
            ws.broadcast_changes(values, field='value')
        tbl.needUpdate = not recalc

    def fml_dependents(self, df):
        rgn_fmls = df['fml'].to_dict()
        independents = {}
        for coord, fml in rgn_fmls.items():
            pairs = set()
            for term in rgn_pattern.findall(fml):
                term = term.replace('$', '')
                if ':' in term:
                    ws_name, cell_rng = f"!{term}".split('!')[-2:]
                    ws_name = f'{ws_name}!' if ws_name else ''
                    rgn_coords = coords_from_range(cell_rng)
                    pairs.update([f'{ws_name}{chr(ord("A") + col - 1)}{row}'
                                  for row in range(rgn_coords['min_row'], rgn_coords['max_row'])
                                  for col in range(rgn_coords['min_col'], rgn_coords['max_col'])
                                  ])
                else:
                    pairs.add(term)
            independents[coord] = pairs
        dependents = collections.defaultdict(set)
        for coord, terms in independents.items():
            for term in terms:
                dependents[term].add(coord)
        df = pd.concat(
            [
                df,
                pd.DataFrame(independents.items(), columns=['coord', 'independents']).set_index('coord'),
                pd.DataFrame(dependents.items(), columns=['coord', 'dependents'], dtype=TABLE_DATA_MAP['dependents']).set_index('coord')
            ],
            axis=1
        )
        return df

    def res_order(self, df):
        independents = dict(df['independents'].to_dict().items())
        pairs = set(
            (key, dep) for key, deps in independents.items()
            if isinstance(deps, set)
            for dep in deps
        )
        term_dep, term_ind = zip(*pairs)
        res_order = [set(term_ind) - set(term_dep)]
        counter = collections.Counter(term_dep)
        while pairs:
            # print(f'For process={len(pairs)}')
            to_batch = [
                (term, dep)
                for term, dep in pairs
                if dep in res_order[-1]
            ]
            counter.subtract([
                term
                for term, dep in to_batch
                if not pairs.remove((term, dep))
            ])
            to_batch = {term for term in counter.keys() if counter[term] == 0}
            [counter.pop(term) for term in to_batch]
            res_order.append(to_batch)
        res_order_df = pd.concat(
            pd.DataFrame(k, index=sorted(res_order[k]), columns=['res_order'], dtype=TABLE_DATA_MAP['res_order'])
            for k in range(len(res_order)
                           )
        )
        df = pd.concat(
            [
                df,
                res_order_df
            ],
            axis=1
        )
        return df

    def _cell_rgn(self, excel_slice: tuple[str, ...] | str) -> list[str]:
        (rmin, cmin),  (rmax, cmax) = map( cell_address, self.data_rng.split(':'))
        rmin, rmax = int(rmin), int(rmax)

        if isinstance(excel_slice, str):
            excel_slice = (excel_slice, )
        cell_rgn = set()
        for slice in excel_slice:
            sheet, slice = tbl_address(slice)
            prefix = '' if sheet in (None, self.parent.title) else f'\'{sheet}\'!'
            slice = slice.upper().replace('$', '')
            linf, lsup = f'{slice}:{slice}'.split(':', 2)[:2]
            if (linf + lsup).isnumeric():
                if not prefix and (rmin <= int(linf) <= rmax) and (rmin <= int(lsup) <= rmax):
                    linf = f'{cmin}{linf}'
                    lsup = f'{cmax}{lsup}'
                else: 
                    continue
            if (linf + lsup).isalpha():
                if not prefix and (cmin <= linf <= cmax) and (cmin <= lsup <= cmax):
                    linf = f'{linf}{rmin}'
                    lsup = f'{lsup}{rmax}'
                else:    
                    continue
            linf_cell = cell_address(linf)
            lsup_cell = cell_address(lsup)
            cell_rgn.update(
                [
                    f'{prefix}{col}{row}' for row, col in itertools.product(
                    [x for x in range(int(linf_cell[0]), int(lsup_cell[0]) + 1)],
                    [chr(x) for x in range(ord(linf_cell[1]), ord(lsup_cell[1]) + 1)])
                ]
            )
        cell_rgn = sorted(cell_rgn, key=lambda x: '{0: >4s}{1}'.format(*cell_address(x)))
        return cell_rgn

    @classmethod
    def excel_table(cls, data:pd.Series):
        return (
            data
            .set_axis(
                pd.MultiIndex.from_tuples(
                    [cell_address(cell)[:2] for cell in data.index],
                    names=['row', 'col']
                )
            )
            .unstack(level=-1)
            .sort_index(key=lambda x: x.str.extract(r'(\d+)', expand=False).astype(int))
        )

    def encoder(self, action: Literal['encode', 'decode'], in_fmls: str | list[str] | pd.Series, df: pd.DataFrame | None = None) -> pd.Series:
        def translate(m, field: Literal['cell', 'code']):
            var = df.loc[m[0].replace('$', ''), field]
            return pass_anchors(m[0], var)

        if b_str := isinstance(in_fmls, str):
            in_fmls = pd.Series([in_fmls])
        if df is None:
            df = self.data
        if action == 'decode':
            df = df.reset_index().set_index('code')
            field = 'cell'
        else:
            field = 'code'
        if (b_lst := isinstance(in_fmls, list)):
            in_fmls = pd.Series(in_fmls)
        out_fmls = in_fmls.str.replace(
            cell_pattern,
            # lambda m: pass_anchors(m[0], df.loc[m[0].replace('$', ''), field]),
            lambda m: translate(m, field),
            regex=True
        )
        if b_lst:
            out_fmls = out_fmls.tolist()
        return out_fmls.iloc[0] if b_str else out_fmls

    def get_formula(self, *excel_slice: tuple[str, ...]):
        cell_rgn = self._cell_rgn(excel_slice)
        coded_fmls = self.data.loc[cell_rgn, 'fml']
        fmls = self.encoder('decode', coded_fmls)
        if len(cell_rgn) == 1:
            return fmls.iloc[0]
        return self.excel_table(fmls)
    
    def offset_rng(self, cells: str | list[str], col_offset: int = 0, row_offset: int = 0, disc_cell: str | None = None) -> str | dict[str, str]:
            if bflag := isinstance(cells, str):
                cells = [cells]

            disc_sht = [None, self.parent.title]
            predicate = lambda x: True
            # Cuando se eliminen celdas, se debe asegurar que el offset no sobrepase los límites de la tabla
            rmin, cmin = 1, ord('A')
            if disc_cell:
                sht, disc_cell = tbl_address(disc_cell)
                row, col = cell_address(disc_cell)
                rmin = int(row)
                cmin = ord(col)
                if sht not in disc_sht:
                    disc_sht = [sht]
                if col_offset == 0 and row_offset:
                    predicate = lambda x: int(cell_address(x)[0]) >= int(cell_address(disc_cell)[0])
                if col_offset and row_offset == 0:
                    predicate = lambda x: ord(cell_address(x)[1]) >= ord(cell_address(disc_cell)[1])
                else:
                    predicate = lambda x: '{0: >4s}{1}'.format(*cell_address(x)) >= '{0: >4s}{1}'.format(*cell_address(disc_cell))
                if col_offset == 0 and row_offset:
                    predicate = lambda x: int(cell_address(x)[0]) >= int(cell_address(disc_cell)[0])
                if col_offset and row_offset == 0:
                    predicate = lambda x: ord(cell_address(x)[1]) >= ord(cell_address(disc_cell)[1])
                else:
                    predicate = lambda x: '{0: >4s}{1}'.format(*cell_address(x)) >= '{0: >4s}{1}'.format(*cell_address(disc_cell))

            try:
                filter_rng, filter_sht, filter_cells = zip(
                    *[
                        (x, *tbl_addr) for x in cells 
                        if (tbl_addr := tbl_address(x)) and tbl_addr[0] in disc_sht 
                    ]
                )
            except ValueError:
                answ = {}
            else:
                ndx = pd.Index(
                    [x for x in itertools.chain(*[y.split(':') for y in filter_cells]) if predicate(x)]
                )
                db = ndx.str.extract(cell_pattern, expand=True).set_index(ndx)

                mask = ~db.row.str.contains('$', regex=False)
                fnc = lambda x: str(max(rmin, int(x.strip('$')) + row_offset))
                db.loc[mask, 'row'] = db.loc[mask, 'row'].apply(fnc)
                
                mask = ~db.col.str.contains('$', regex=False)
                fnc = lambda x: chr(max(cmin, ord(x) + col_offset))
                db.loc[mask, 'col'] = db.loc[mask, 'col'].apply(fnc)

                db['cell'] = db.col + db.row
                cells_map = db.cell.to_dict()
                values = [':'.join(map(lambda x: cells_map.get(x, x), key.split(':'))) for key in filter_cells]
                values = [(f"'{sht}'!" if sht else '') + value for sht, value in zip(filter_sht, values)]
                answ = dict(zip(filter_rng, values))
            return answ.get(cells[0], cells[0]) if bflag else answ
    
    def all_dependents(self, cells):
        reduce = lambda items: functools.reduce(lambda t, e: t.union(e) or t, items, set())
        ws = self.parent
        wb = ws.parent
        tbl = self
        # Se limita el alcance a las cells solicitadas
        df = tbl.data.loc[cells]
        # Se da la posibilidad de que en las cells puedan venir parámetros
        codes = ['!'.join(f"'{ws.id}'!{code}".split('!')[-2:]) for code in df.code]
        fmls = []
        mask = ~df.dependents.isna()
        dep_df = df.loc[mask, ['code', 'dependents']]
        dep_pairs = [(f"'{ws.id}'!{code}", f"'{ws.id}'!{x}") for _, (code, dep) in dep_df.iterrows() for x in dep]
        fml_df = pd.DataFrame(dep_pairs, columns=['code', 'dependent'])
        fmls.append(fml_df)

        external_links = wb.links.keys() & set(f"'{ws.id}'!{x}" for x in df.loc[cells].code)
        etables = reduce([wb.links[code] for code in external_links])
        for etbl in etables:
            sht_id, tbl_id = tbl_address(etbl)
            df = wb['#' + sht_id]['#' + tbl_id].data

            dep_df = df.loc[df.code.isin(codes), ['code', 'dependents']]
            dep_pairs = [(code, f"'{sht_id}'!{x}") for _, (code, dep) in dep_df.iterrows() for x in dep]
            fml_df = pd.DataFrame(dep_pairs, columns=['code', 'dependent'])
            fmls.append(fml_df)
        fmls = (
            pd.concat(fmls)
            .drop_duplicates()
            .set_index('code')
            .sort_index()
        )
        return fmls

    def __getitem__(self, excel_slice: str|list[str]) -> pd.DataFrame:
        cell_rgn = self._cell_rgn(excel_slice)
        df = self.excel_table(self.data.loc[cell_rgn, 'value'])
        return df

    def __setitem__(self, excel_slice, value):
        cell_rgn = self._cell_rgn(excel_slice)
        self.data.loc[cell_rgn, 'value'] = np.array(value).flatten()
        codes = self.encoder('encode', cell_rgn)
        self.changed.extend(codes)
        codes = self.encoder('encode', cell_rgn)
        self.changed.extend(codes)

    def __contains__(self, item):
        row, col, sh_name = (cell_address(item) + (None, ))[:3]
        sh_name = sh_name or self.parent.title
        if (bflag := sh_name == self.parent.title):
            (rmin, cmin),  (rmax, cmax) = map( lambda x: (int((tpl:=cell_address(x))[0]), f'{tpl[1]: >2s}'), self.data_rng.split(':'))
            bflag = (rmin <= int(row) <= rmax) and (cmin <= f'{col.upper(): >2s}' <= cmax)
        return bflag

    def minimun_table(self):
        cell_rgn = self.cells_in_data_rng(self.data.index.tolist())
        mask = (self.data.index.isin(cell_rgn)) & (self.data.value != 0)
        excel_slice = self.data.loc[mask, :].index.tolist()
        df = self[excel_slice].fillna(0)
        return df

    def _repr_html_(self):
        all_cells = self._cell_rgn(self.data_rng)
        t = pd.Series('', index=all_cells)
        t[self.data.index] = self.data.value
        data = self.excel_table(t)
        return data._repr_html_()

def test_Form2517():
        filename = r"C:\Users\agmontesb\Documents\DIAN\Renta2023\Reporte_Conciliación_Fiscal_F2517V6_AG2023_v1.0.1-2024\Reporte_Conciliación_Fiscal_F2517V6_AG2023_v1.0.1-2024.xlsm"
        wb = px.load_workbook(filename)

        excel_wb = ExcelWorkbook('Form2517')

        ws_name = "H2 (ESF - Patrimonio)"
        ws = wb[ws_name]
        wsheet = excel_wb.create_worksheet(ws_name)

        # Estado de Situación Financiera
        ws_range = "G9:K193"
        fmls, values = data_in_range(ws, ws_range)
        esf_tbl = ExcelTable(wsheet, 'esf_tbl', ws_range, fmls, values, recalc=True)
        esf = esf_tbl.minimun_table()

        # Patrimonio
        ws_range = "G196:K220"
        fmls, values = data_in_range(ws, ws_range)
        pat_tbl = ExcelTable(wsheet, 'pat_tbl', ws_range, fmls, values, recalc=True)
        pat = pat_tbl.minimun_table()

        ws_name = "H3 (ERI - Renta Liquida)"
        ws = wb[ws_name]
        wsheet = excel_wb.create_worksheet(ws_name)

        # Renta líquida cedular
        ws_range = "H376:N457"      # "H10:L10"
        fmls, values = data_in_range(ws, ws_range)
        # values['L10'] = 33182000
        # values['L25'] = 13287242
        # values['M6'] = 42412
        rlc_tbl = ExcelTable(wsheet, 'rlc_tbl', ws_range, fmls, values, recalc=True)
        wsheet.parameters(M6 = 42414)
        # dmy = rlc_tbl.get_formula('L412', 'H420')
        rlc = rlc_tbl.minimun_table()

        # Estado de Resultados
        ws_range = "H9:L375"      # "H10:L10"
        fmls, values = data_in_range(ws, ws_range)
        edr_tbl = ExcelTable(wsheet, 'edr_tbl', ws_range, fmls, values, recalc=True)
        edr = edr_tbl.minimun_table()


        ws_name = 'H7 (Resumen ESF-ERI)'
        ws = wb[ws_name]
        wsheet = excel_wb.create_worksheet(ws_name)

        ws_range = 'G11:I93'
        fmls, values = data_in_range(ws, ws_range)
        res_tbl = ExcelTable(wsheet, 'res_tbl', ws_range, fmls, values, recalc=True)
        res = res_tbl.minimun_table()

        wb.close()
        return excel_wb

def test_workbook():
        filename = r"C:\Users\agmontesb\Downloads\excel_module_test.xlsx"
        wb = px.load_workbook(filename)

        excel_wb = ExcelWorkbook('excel_module_test')

        ws_name = "No links, No parameters"
        ws = wb[ws_name]
        wsheet = excel_wb.create_worksheet(ws_name)

        # Tabla 1
        ws_range = "G4:I9"
        fmls, values = data_in_range(ws, ws_range)
        sh1_tbl1 = ExcelTable(wsheet, 'sh1_tbl1', ws_range, fmls, values, recalc=True)
        m_sh1_tbl1 = sh1_tbl1.minimun_table()

        # Tabla 2
        ws_range = "G13:H15"
        fmls, values = data_in_range(ws, ws_range)
        sh1_tbl2 = ExcelTable(wsheet, 'sh1_tbl2', ws_range, fmls, values, recalc=True)
        m_sh1_tbl2 = sh1_tbl2.minimun_table()

        ws_name = "Parameters and inner links"
        ws = wb[ws_name]
        wsheet = excel_wb.create_worksheet(ws_name)

        # Tabla 1
        ws_range = "F4:H9"
        fmls, values = data_in_range(ws, ws_range)
        sh2_tbl1 = ExcelTable(wsheet, 'sh2_tbl1', ws_range, fmls, values, recalc=True)
        m_sh2_tbl1 = sh2_tbl1.minimun_table()

        # Tabla 2
        ws_range = "F13:H17"
        fmls, values = data_in_range(ws, ws_range)
        sh2_tbl2 = ExcelTable(wsheet, 'sh2_tbl2', ws_range, fmls, values, recalc=True)
        m_sh2_tbl2 = sh2_tbl2.minimun_table()

        ws_name = "Outer links, outer parameter"
        ws = wb[ws_name]
        wsheet = excel_wb.create_worksheet(ws_name)

        # Tabla 1
        ws_range = "F3:H8"
        fmls, values = data_in_range(ws, ws_range)
        sh3_tbl1 = ExcelTable(wsheet, 'sh3_tbl1', ws_range, fmls, values, recalc=True)
        m_sh3_tbl1 = sh3_tbl1.minimun_table()

        wb.close()

        return excel_wb

    

if __name__ == '__main__':
        import openpyxl as px
        excel_wb = test_workbook()

        pass