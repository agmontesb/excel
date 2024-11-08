import sys
import os
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import pytest
import openpyxl as px
import pandas as pd
import itertools
import functools

from excel_workbook import (
    TABLE_DATA_MAP, ExcelWorkbook, ExcelTable, XlErrors, 
    cell_address, cell_pattern, 
    tbl_address, rgn_pattern,
    CIRCULAR_REF, link_pattern
    )
from utilities import TableComparator, tbl_data
from fixtures import static_workbook as base_workbook


class TestBaseWorkbook:

    def test_base_workbook(self, base_workbook):

        assert base_workbook.title == 'excel_module_test'
        assert base_workbook.parent is None

        sheet_names = ['No links, No parameters', 'Parameters and inner links', 'Outer links, outer parameter']
        assert base_workbook.sheetnames == sheet_names

        sheet_ids = ['sheet', 'sheet1', 'sheet2']
        assert sheet_ids == [sheet.id for sheet in base_workbook.sheets]

        assert all(
            base_workbook.index(base_workbook.sheets[k]) == k 
            for k in range(len(base_workbook.sheets))
            )
        
        assert all(base_workbook[f'#{id}'].title == sheet_name for id, sheet_name in zip(sheet_ids, sheet_names))
        assert all(base_workbook[sheet_name].id == id for id, sheet_name in zip(sheet_ids, sheet_names))

    def test_base_worksheet(self, base_workbook):
        sheet = base_workbook['No links, No parameters']
        assert sheet.title == 'No links, No parameters'
        assert sheet.parent is base_workbook
        assert sheet.id == 'sheet'

        table_names = ['sht1_tbl2', 'sht1_tbl2']
        assert sheet.tablenames == table_names
        assert sheet.tables == [sheet[sheet_name] for sheet_name in table_names]

        assert all(sheet.index(sheet.tables[k]) == k for k in range(len(sheet.tables)))

        sheet_ids = [sheet.id for sheet in sheet.tables]
        assert all(sheet[f'#{id}'].title == table_name for id, table_name in zip(sheet_ids, table_names))

    def test_base_table(self, base_workbook):
        table = base_workbook['No links, No parameters']['sht1_tbl1']
        assert table.title == 'sht1_tbl1'
        assert table.parent is base_workbook['No links, No parameters']
        assert table.id == 'A'

        assert table.data_rng == 'G4:I9'
        assert table.needUpdate is False
        assert table.changed == []

        assert table.parent.index(table) == 0
        assert table.parent.objectnames() == ['sht1_tbl1', 'sht1_tbl2']

    def test_ws_parameters(self, base_workbook):
        ws = base_workbook['Parameters and inner links']
        assert ws.parameters() == ['G2']
        assert ws.parameters('G2') == [0]

        associated_tables = ws.associated_table('G2', scope='parameter')
        assert sorted(tbl.title for tbl in associated_tables) == ['sht2_tbl2', 'sht3_tbl1']

        param_code = ws.parameter_code('G2')
        dependents = [
            tbl.encoder('decode', tbl.get_cells_to_calc([param_code]), df=tbl.data)
            for tbl in associated_tables
            ]
        old_values = [
            tbl.data.loc[dependent, 'value'].tolist() for 
            tbl, dependent in zip(associated_tables, dependents)
            ]
        
        ws.parameters(G2=42420)

        new_values = [
            tbl.data.loc[dependent, 'value'].tolist() for 
            tbl, dependent in zip(associated_tables, dependents)
            ]
        
        assert all(42420 in value for value in new_values), 'Change parameter not reported'
        assert all(old != new for old, new in zip(old_values, new_values)), 'Change parameter not applied'


    def test_ws_cell_values(self, base_workbook):
        ws = base_workbook['Outer links, outer parameter']
        tbl = ws['sht3_tbl1']

        links = tbl.links()
        links.append('A1')  # Not in the table range nor in the links
        links, values = ws.cell_values(links)

        # All input cells are sheet discriminated
        assert all('!' in x for x in links)
        # For cells not in the table range or in the links, the value is None
        assert [tbl_address(cell)[-1] for cell, value in zip(links, values) if value is None] == ['A1']


class TestHelperMethods:

    @pytest.mark.parametrize("cells, row_offset, col_offset, disc_cell, answer, msg", [
        ('A1', 1, 0, None, 'A2', 'Offset row'),
        ('A1',1, 1, None,'B2', 'Offset row and col'),
        ('$A1',1, 1, None,'$A2', 'Offset row and col'),
        ('A$1',1, 1, None,'B$1', 'Offset row and col'),
        ('$A$1',1, 1, None,'$A$1', 'Offset row and col'),
        # Offset range
        ('A1:C5', 0, 1, None,'B1:D5', 'Offset col'),
        # The disc_cell parameter: Cells be offseted if cell > = disc_cell
        # The disc_cell worksheet name if not present to be consider as tbl parent worksheet title
        ('D4', 1, 0, 'A5', 'D4', 'D4 < A5 => D4'),
        ('B5', 1, 0, 'A5', 'B6', 'B5 > A5 => B6'),
        # The disc_cell worksheet name offset only cells in the same worksheet
        ("'sheet12'!C5", 1, 0, "'sheet12'!A5","'sheet12'!C6", ""),
        ('C5', 1, 0, "'sheet12'!A5",'C5', ""),
    ])
    def test_offset_rng_blk1(self, cells, row_offset, col_offset, disc_cell, answer, msg, base_workbook):
        assert ExcelTable.offset_rng(cells, col_offset, row_offset, disc_cell) == answer, msg


    def test_offset_rng_blk2(self, base_workbook):
        ws = base_workbook['Parameters and inner links']
        tbl = ws['sht2_tbl1']

        # Equivalentes
        r1 = ExcelTable.offset_rng('B5', row_offset=1, disc_cell="A5", tbl=tbl)
        r2 = ExcelTable.offset_rng('B5', row_offset=1, disc_cell=f"'{ws.title}'!A5", tbl=tbl)
        r3 = ExcelTable.offset_rng(f"'{ws.title}'!B5", row_offset=1, disc_cell="A5", tbl=tbl)
        r4 = ExcelTable.offset_rng(f"'{ws.title}'!B5", row_offset=1, disc_cell=f"'{ws.title}'!A5", tbl=tbl)
        assert r1 == r2 == r3.split('!')[-1] == r4.split('!')[-1], f'{r1=} != {r2=} != {r3=} != {r4=}'

        # Offset list of cells = offset each cell
        predicate = lambda x, disc_cell: '{0: >4s}{1}'.format(*cell_address(x)) >= '{0: >4s}{1}'.format(*cell_address(disc_cell))
        cells = ['A1', 'B10', 'C5', 'D4']
        kwargs = dict(row_offset=1, col_offset=1, disc_cell='A5', tbl=tbl)
        fnc = lambda x: ExcelTable.offset_rng(x, **kwargs)
        cells_map = ExcelTable.offset_rng(cells, **kwargs)
        assert all(fnc(key) == value for key, value in cells_map.items()), 'Offset list of cells'
        assert len(cells) >= len(cells_map), 'Offset list of cells'
        assert all(predicate(x, kwargs['disc_cell']) for x in (set(cells) - cells_map.keys()))


class TestErrorHandling:

    @pytest.mark.parametrize("error_type, fml", map(lambda x: pytest.param(*x, id=f'ErrorType - {x[0]}'), [
        (XlErrors.DIV_ZERO_ERROR, '=1/0'),
        (XlErrors.VALUE_ERROR, '=1 + "uno"'),
        (XlErrors.NAME_ERROR, '=1 + dos'),
    ]))
    def test_error_types(self, base_workbook, error_type, fml):
        tbl = base_workbook['Parameters and inner links']['sht2_tbl1']
        assert tbl.evaluate(fml) == error_type

    @pytest.mark.parametrize("xl_error, error_origen, fml", map(lambda x: pytest.param(*x, id=f'Propagate - {x[0]}'), [
        (XlErrors.DIV_ZERO_ERROR, 'G6', '=1/0'),
        (XlErrors.VALUE_ERROR, 'G6', '=1 + "uno"'),
        (XlErrors.NAME_ERROR, 'G6', '=1 + dos'),
    ]))
    def test_error_propagation(self, base_workbook, xl_error, error_origen, fml):
        tbl = base_workbook.sheets[1].tables[0]

        # En un libro sin errores
        df1 = base_workbook.data
        mask = df1.value.apply(lambda x: isinstance(x, XlErrors)) & ~df1.code.str.contains('!')
        assert not mask.any()

        # Al producirse un error en el cálculo de una celda, 
        tbl.set_records({error_origen: fml}, field='fml')
        tbl.recalculate(recalc=True)
        df = base_workbook.data

        mask = df.value.apply(lambda x: isinstance(x, XlErrors)) & ~df.code.str.contains('!')
        assert mask.any()

        # el error se propaga a todas las celdas relacionadas con la celda afectada.
        error_cells, error_codes = zip(*(df.loc[mask].tbl_id.str[0:-1] + df.loc[mask].code).items())
        tbl_codes = df.loc[mask].tbl_id
        assert len(error_cells) >= 1 and error_origen in error_cells

        # Las celdas diferentes a la celda afectada, son celdas de fórmulas que para su cálculo 
        # dependen del valor de la celda afectada.
        all_deps = tbl.parent.all_dependents({tbl:{ExcelTable.encoder('encode', error_origen, df=tbl.data)}}, with_links=True)
        dependents = [f"{prefix}{code}" for tbl, codes in all_deps.items() for code in codes if (prefix:=f"'{tbl.parent.id}'!" if '!' not in code else '')]

        assert not set(error_codes) ^ set(dependents)

        # Cualquier cálculo que involucre a una celda con un error, tiene como resultado 
        # un error del mismo tipo.
        for cell, code in zip(error_cells, tbl_codes):
            sht_id, tbl_id = tbl_address(code)
            tbl = base_workbook['#' + sht_id]['#' + tbl_id]
            assert tbl.evaluate(f'={cell} + 25') is xl_error

        # Se elimina el error eliminando el origen del mismo.
        tbl.set_records({error_origen: 100}, field='value')
        tbl.recalculate(recalc=True)
        df = base_workbook.data

        mask = df.value.apply(lambda x: isinstance(x, XlErrors)) & ~df.code.str.contains('!')
        assert mask.any()


class TestModTables:

    @pytest.mark.parametrize("ws_name, cell_slice, answer", [
        ('Parameters and inner links', 'G', dict(unchanged=[], zoomed_in=['sht2_tbl1', 'sht2_tbl2'], displaced=[])),
        ('Parameters and inner links', 'F', dict(unchanged=[], zoomed_in=['sht2_tbl1', 'sht2_tbl2'], displaced=[])),
    ])
    def test_delete(self, base_workbook, ws_name, cell_slice, answer):

        tables1 = {tbl.title: tbl.data_rng for ws in base_workbook.sheets for tbl in ws.tables}
        ws = base_workbook[ws_name]

        tbl = base_workbook['Outer links, outer parameter']['sht3_tbl1']
        tbl.set_records({'E8': "=+'Parameters and inner links'!F5"}, field='fml')
        tbl.recalculate(recalc=True)

        # Códigos a borrar en el libro de trabajo
        all_codes = {}
        [all_codes.setdefault(tbl, []).append(ExcelTable.encoder('encode', cell, df=tbl.data))
            for tbl in ws.tables
            for cell in tbl._cell_rgn(cell_slice)
            if cell in tbl.data.index
        ]
        del_codes = [code for codes in all_codes.values() for code in codes]

        all_deps = ws.all_dependents(all_codes, with_links=True)

        tdf1 = TableComparator(base_workbook.data.assign(case=1))
        ws.delete(cell_slice)
        tdf2 = TableComparator(base_workbook.data.assign(case=2))

        tables2 = {tbl.title: tbl.data_rng for ws in base_workbook.sheets for tbl in ws.tables}

        assert not (tables1.keys() ^ tables2.keys())  # En este test no se ha eliminado ninguna tabla

        # Modificación del rango de datos en las tablas de la hoja de trabajo 
        # donde se realizó la operación de eliminación
        tbl_clasifier_keys = [
            'unchanged',   # tbls inalteradas, se insertan filas/cols en rangos que no hacen parte de la tabla.
            'zoomed_in',      # tbls cuyo rango de datos se amplia por la operación de inserción
            'displaced'    # tbls cuyo rango de datos cambia de ubicación.
        ]
        tbl_clasifier = lambda key, tbl1, tbl2: len(set(f'{tbl1[key]}:{tbl2[key]}'.split(':'))) - 2

        tbl_clasifier_map = {}
        {
            tbl_clasifier_map.setdefault(key, []).append(tbl_title) 
            for tbl_title in tables1 
            if (key:=tbl_clasifier_keys[tbl_clasifier(tbl_title, tables1, tables2)])
        }

        if len(tbl_clasifier_map) == 1 and 'unchanged' in tbl_clasifier_map:
            # No se produjo cambio
            return

        # Esta operación afecta solo la estructura de la hoja donde se lleva a cabo.
        ws_tables = set(ws.tablenames)

        assert len(
            [
                key for key, items in tbl_clasifier_map.items() 
                if set(items) & ws_tables
            ]
        ) <= 3

        # Deja inalterada la estructura de las otras hojas del libro. 
        assert ['unchanged'] == [
            key for key, items in tbl_clasifier_map.items() 
            if set(items) - ws_tables
        ]

        # CAMBIOS EN LA HOJA DONDE SE LLEVA A CABO LA OPERACIÓN

        # Luego de la operación, no existen los códigos a borrar en el libro de trabajo.
        mask = tdf2.df.code.isin(del_codes) & tdf2.df.tbl_id.str.startswith(f"'{ws.id}'!")
        assert tdf2.df.loc[mask].empty

        # Los cambios en las hojas diferentes a la afectada se producen en las fórmulas enlazadas
        # con los códigos a borrar.
        all_tbl_ids = set(tdf1.df.tbl_id)
        tbls_id_deps = set(f"'{tbl.parent.id}'!{tbl.id}" for tbl in all_deps)
        unchanged_tbls = all_tbl_ids.difference(tbls_id_deps)

        diff = tdf1.symmetric_difference(tdf2, fields=['cell', *TABLE_DATA_MAP.keys()])
        
        cslice = f'{cell_slice}:{cell_slice}'.split(':', 2)[-2:]
        kwargs = dict(row_offset=0, col_offset=0)
        isRow = int(cslice[0].isnumeric())
        fnc = [ord, int][isRow]
        key = ['col_offset', 'row_offset'][isRow]
        kwargs[key] = fnc(cslice[1]) - fnc(cslice[0]) + 1
        kwargs['disc_cell'] = f'A{cslice[0]}' if isRow else f'{cslice[0]}1'

        mask = 	diff.tbl_id.str.startswith(f"'{ws.id}'!") & ~diff.code.str.contains('!')
        diff_ws = diff.loc[mask]
        diff_ws.loc[diff_ws.case == 2, 'cell'] = diff_ws.loc[diff_ws.case == 2].cell.apply(lambda x: ExcelTable.offset_rng(x, **kwargs))
        diff_ws = diff_ws.sort_values(by=['cell'])
        diff_ws = diff_ws.drop_duplicates(subset=['cell', *TABLE_DATA_MAP.keys()], keep=False)

        unique_cells = [key for key, bflag in list((diff_ws.cell.value_counts() == 1).items()) if bflag]
        # The unique cells are a subset of the deledted cells (del_codes) because is posible that
        # in del_codes are empty cells that are not present in the data. 
        assert set(diff_ws.loc[diff_ws.cell.isin(unique_cells)].code).issubset(del_codes)
        # The del_codes are the only ones with case == 1 because after that the code no longer exists
        assert (diff_ws.loc[diff_ws.cell.isin(unique_cells)].case == 1).all()
        # We don't need the deleted cells anymore.
        diff_ws = diff_ws.loc[~(diff_ws.cell.isin(unique_cells))]

        # The formulas are fully cualified (i.e. they have the sheet name) for efect of concatenation
        # in next steps.
        fnc = lambda x: cell_pattern.sub(lambda m: '!'.join(f"'{ws.id}'!{m[0]}".split('!')[-2:]), x)
        diff_ws.loc[:, 'fml'] = diff_ws.fml.apply(fnc)

        # Differences in other tables not present in the worksheet where the operation 
        # was performed.
        mask = 	~diff.tbl_id.str.startswith(f"'{ws.id}'!") & ~diff.code.str.contains('!')
        diff_other = diff.loc[mask]

        assert diff_other.loc[diff_other.tbl_id.isin(unchanged_tbls)].empty
        assert set(diff_other.tbl_id).issubset(tbls_id_deps)

        # From here, the tests can be done in diff_ws and diff_other, so we concatenate 
        # then to make it at the same time.
        diff_all = pd.concat([diff_ws, diff_other])

        # full cualified codes for the deleted cells.
        del_fcodes = set(f"'{ws.id}'!{code}" for code in del_codes)

        # tres tipos de records:
        # 1 - Con valor '#REF!', que tienen su origen en celdas que tienen referencia 
        # directa a las celdas eliminadas

        mask  = diff_all.fml.str.contains("'sheet1'!ZZ0")
        cells = set(diff_all.loc[mask].cell)
        rec_type1 = diff_all.loc[diff_all.cell.isin(cells)]
        # Todas las fórmulas en rec_type1 hacen referencia directa a las celdas borradas, 
        # ya sea en referencia de celda o referencia de rango en ambos extremos del rango,   
        mask = rec_type1.fml.map(lambda x: bool(set(link_pattern.findall(x)) & del_fcodes))
        # antes de la operación (case = 1)
        assert (rec_type1.loc[mask].case == 1).all()

        # Se eliminan las rec_type1 que ya no necesitaremos
        diff_all = diff_all.drop(index=rec_type1.index)

        # 2 - Con valor '#REF!', en celdas que tienen referencia a las celdas del tipo 1.
        if not diff_all.empty:
            mask  = diff_all.value.isin([XlErrors.REF_ERROR])
            cells = set(diff_all.loc[mask].cell)
            rec_type2 = diff_all.loc[diff_all.cell.isin(cells)]
            # cells con referencia a rango en que uno de los extremos hace referencia a los códigos 
            # a borrar
            mask = (rec_type2.case == 1) & rec_type2.fml.map(lambda x: len(set(link_pattern.findall(x)) & del_fcodes) % 2 == 1)
            cells1 = set(rec_type2.loc[mask].cell)
            # cells sin que dependen de las celdas en rec_type1.
            mask = (rec_type2.case == 1) & ~rec_type2.fml.map(lambda x: len(set(link_pattern.findall(x)) & del_fcodes))
            cells2 = set(rec_type2.loc[mask].cell)
            # antes de la operación (case = 1)
            assert cells == cells1 | cells2
            # Se eliminan las rec_type1 que ya no necesitaremos
            diff_all = diff_all.drop(index=rec_type2.index)

        # 3 - Con valor diferente a '#REF!', cuyo valor depende de las nuevas celdas renombradas.
        assert (~diff_all.value.isin([XlErrors.REF_ERROR])).all()


    @pytest.mark.parametrize("ws_name, cell_slice, answer", [
        ('Parameters and inner links', 'F', dict(unchanged=[], zoomed=['sht2_tbl1', 'sht2_tbl2'], displaced=[])),
        ('Parameters and inner links', 'M', dict(unchanged=['sht2_tbl1', 'sht2_tbl2'], zoomed=[], displaced=[])),
        ('Parameters and inner links', '40', dict(unchanged=['sht2_tbl1', 'sht2_tbl2'], zoomed=[], displaced=[])),
        ('Parameters and inner links', '15', dict(unchanged=['sht2_tbl1'], zoomed=['sht2_tbl2'], displaced=[])),
        ('Parameters and inner links', '10', dict(unchanged=['sht2_tbl1'], zoomed=[], displaced=['sht2_tbl2'])),
        ('Parameters and inner links', '6', dict(unchanged=[], zoomed=['sht2_tbl1'], displaced=['sht2_tbl2'])),
        ('Parameters and inner links', '1', dict(unchanged=[], zoomed=[], displaced=['sht2_tbl1', 'sht2_tbl2'])),
    ])
    def test_insert(self, base_workbook, ws_name, cell_slice, answer):
        tbl_map_keys = list(answer.keys())
        ws = base_workbook[ws_name]
        tbl_range1 = {tbl.title: tbl.data_rng for tbl in ws.tables}

        tdf1 = TableComparator(base_workbook.data.assign(case=1))
        ws.insert(cell_slice)
        tdf2 = TableComparator(base_workbook.data.assign(case=2))

        tbl_range2 = {tbl.title: tbl.data_rng for tbl in ws.tables}

        # La operación de inserción afecta el rango de datos de las tablas en la 
        # hoja de trabajo donde se hace la inserción así:
        tbl_map = dict(
            unchanged=[],   # tbls inalteradas, se insertan filas/cols en rangos que no hacen parte de la tabla.
            zoomed=[],      # tbls cuyo rango de datos se amplia por la operación de inserción
            displaced=[]    # tbls cuyo rango de datos cambia de ubicación.
        )
        fnc = lambda tid: len(set(f'{tbl_range1[tid]}:{tbl_range2[tid]}'.split(':'))) - 2
        pairs = [(fnc(tbl.title), tbl.title) for tbl in ws.tables]
        [tbl_map[key].append(tbl_title) for k, tbl_title in pairs if (key:=tbl_map_keys[k])]
        assert tbl_map == answer

        if not (tbl_map['zoomed'] or tbl_map['displaced']):
            # No se produjo cambio
            return
        # Se produjo cambio:
        # El cambio en los rango de datos refleja un cambio en la estructura de datos..
        changes = tdf1.symmetric_difference(tdf2, fields=['cell', *TABLE_DATA_MAP.keys()])
        assert not changes.empty
        # Los cambios no se producen en los campos (fields) de los registros de las celdas.
        assert changes.drop_duplicates(subset=TABLE_DATA_MAP.keys(), keep=False).empty

        # Los cambios se producen en las etiquetas (campo 'cell') de las celdas por lo que
        # afectan los enlaces externos de las celdas de la worksheet donde se hace la inserción.
        links = changes.loc[changes.code.str.contains('!')]
        assert links.code.str.startswith(f"'{ws.id}'!").all()

        records = (
            changes.loc[~changes.code.str.contains('!')]
            .sort_values(by=['code', 'case'])
            .loc[:, ['cell', 'code', 'tbl_id']]
            .groupby(by='code')
            .agg({'cell': lambda x: x.tolist(), 'tbl_id': lambda x: x.tolist()[0]})
        )
        # Y esto se debe a los cambios que se dan en los registros de las tablas afectadas.
        assert set(records.tbl_id) == set(f"'{ws.id}'!{ws[tbl_name].id}" for tbl_name in tbl_map['zoomed'] + tbl_map['displaced'])
        assert set(records.tbl_id).issubset(f"'{ws.id}'!{tbl.id}" for tbl in ws.tables)
        # El cambio se da en las etiquetas (campo 'cell') de las celdas y se debe
        # al desplazamiento (offset_rgn) de las etiquetas afectadas por la inserción
        cslice = f'{cell_slice}:{cell_slice}'.split(':', 2)[-2:]
        kwargs = dict(row_offset=0, col_offset=0)
        isRow = int(cslice[0].isnumeric())
        fnc = [ord, int][isRow]
        key = ['col_offset', 'row_offset'][isRow]
        kwargs[key] = fnc(cslice[1]) - fnc(cslice[0]) + 1
        kwargs['disc_cell'] = f'A{cslice[0]}' if isRow else f'{cslice[0]}1'

        bflags = [ExcelTable.offset_rng(x, **kwargs) == y for x, y in records.cell]
        assert all(bflags)

        # No se crean registros para las celdas insertadas.
        assert not any(
            tbl.data.index.isin(tbl._cell_rgn(cell_slice)).any() 
            for tbl in map(lambda x: ws[x], tbl_map['zoomed'])
        )

    def test_insert_rows(self, base_workbook):
        # Dado un workbook, luego de la inserción de una(s) fila(s)/columna(s) se verifica que 
        # los cambios existentes en el workbook se deben al desplazamiento de las etiquetas (cells)
        # en la worksheet donde se realizó la inserción.
        wb = base_workbook
        ws = wb['Parameters and inner links']
        tbl = ws['sht2_tbl1']

        tbl.set_records({'F3': '=F6', 'G3': '=G9'}, field='fml')
        tbl.recalculate(recalc=True)

        ins_slice = '6'
        ins_rng = [f"'{ws.title}'!A6"]

        all_dep = pd.concat([
            tbl.direct_dependents(cells)
            for tbl in ws.tables
            if (cells := tbl.cells_in_data_rng(tbl.data.index.tolist()))
        ])

        codes = set(all_dep.index)
        dep_parts = all_dep.dependent.apply(lambda x: cell_pattern.match(x).groups()).tolist()
        dep_df = (
            pd.DataFrame([(f"'{sht}'!{tbl}", f'{tbl}{num}') for sht, tbl, num in dep_parts], columns=['tbl', 'code'])
            .set_index('tbl')
        )

        # En fmls se tienen todas las celdas codificadas del workbook.

        fmls = []
        for tbl_path, codes_df in dep_df.groupby(level=0):
            sht_id, tbl_id = tbl_address(tbl_path)
            tbl = wb['#' + sht_id]['#' + tbl_id]
            codes = codes_df.code.to_list()
            fml_df = (
                tbl.data.loc[tbl.data.code.isin(codes), ['fml', 'code']]
                .set_index('code')
                .rename(index= lambda x: f"'{sht_id}'!{x}")
            )
            fmls.append(fml_df)

        fmls = pd.concat(fmls).drop_duplicates()

        tbl_rngs = [(tbl.id, tbl.data_rng) for tbl in ws.tables]

        # En fmls1 se tienen todas las fórmulas del workbook traducidas según las 
        # etiquetas (cells) exisentes antes de la inserción.
        fmls1 = []
        for tbl_code, fml in fmls.fml.items():
            sht_id, tbl_id, _ = cell_pattern.match(tbl_code).groups()
            ltbl = wb['#' + sht_id]['#' + tbl_id]
            fmls1.append(ltbl.encoder('decode', fml, df=ltbl.data))

        # Insert a row
        ws.insert(ins_slice)

        # En fmls2 se tienen todas las fórmulas del workbook traducidas según las 
        # etiquetas (cells) existentes después de la inserción.
        fmls2 = []
        for tbl_code, fml in fmls.fml.items():
            sht_id, tbl_id, _ = cell_pattern.match(tbl_code).groups()
            ltbl = wb['#' + sht_id]['#' + tbl_id]
            fmls2.append(ltbl.encoder('decode', fml, df=ltbl.data))

        # Si existen diferencias entre las fórmulas traducidas, se debe 
        # al desplazamiento de las etiquetas (cells) en el workbook.
        flags = []
        for tbl_code, *fml_pair in zip(fmls.index, fmls1, fmls2):
            sht_id, tbl_id, _ = cell_pattern.match(tbl_code).groups()
            ltbl = wb['#' + sht_id]['#' + tbl_id]

            lst1, lst2 = map(rgn_pattern.findall, fml_pair)

            flags.extend(
                [ltbl.offset_rng(x, row_offset=1, disc_cell=ins_rng[0], tbl=ltbl) == y for x, y in zip(lst1, lst2) if x != y]
            )
        assert all(flags)

        # Cambio en el data_rng de la tabla en extremos de rango por debajo de la fila insertada.
        # assert all(ws[title].data_rng == ExcelTable.offset_rng(data_rng, row_offset=1, disc_cell=ins_rng[0], tbl=ws[title]) for title, data_rng in tbl_rngs)

        pass


class TestSetRecordsTable:

    @pytest.mark.parametrize("ws_name, tbl_name, tbl_range", [
        ('No links, No parameters', 'sht1_tbl1', 'F3:I9'),
        ('No links, No parameters', 'sht1_tbl2', 'F12:H15'),
        ('Parameters and inner links', 'sht2_tbl1', 'E3:H9'),
        ('Parameters and inner links', 'sht2_tbl2', 'E12:H17'),
        ('Outer links, outer parameter', 'sht3_tbl1', 'E3:H8'),
    ])
    def test_table_creation(self, ws_name, tbl_name, tbl_range, base_workbook):
            sht_tbl = base_workbook[ws_name][tbl_name]
            base_df = sht_tbl.data.copy()
            sht_tbl.data = ExcelTable.create_dataframe()
            fmls, values, tblv = tbl_data(tbl_name, tbl_range)
            sht_tbl.set_records(fmls, field='fml')
            sht_tbl.set_records(values, field='value')
            sht_tbl.recalculate(recalc=True)
            # Se verifica que se reconstuyen los valores de la tabla
            df = sht_tbl.data
            assert (df.loc[tblv.index].value.map(str) == tblv).all()
            assert (
                TableComparator(base_df)
                .symmetric_difference(
                    TableComparator(df),
                    fields=['res_order', 'value']
                )
            ).empty


    @pytest.mark.parametrize("set_up, values, answer", [
        (dict(G3='=10+F3', H3='=10+G3', F3='=10+H3'), dict(G6='=G3+100'),
        {'F3': '0', 'G3': '', 'H3': ''}
        ),
        (dict(F3='=10',G3='=10+F3', H3='=10+G3', G6='=G3+100'), dict(F3='10+H3'),
        {'F3': '0', 'G3': '20', 'H3': '30'}
        ),
        (dict(F3='=10+F3'), None,  # Referencia a si misma
        {'F3': '0'}
        ),
        (dict(F3='=10+H3',G3='=10+F3', H3='=10+G3'), None,  # Conjunto de celdas con referencia circular.
        {'F3': '', 'G3': '', 'H3': '0'}
        ),
        (dict(F3='=10',G3='=10+F3', H3='=10+G3'), dict(F3='10+H3'),  # Fórmula que cierra la referencia circular.
        {'F3': '0', 'G3': '20', 'H3': '30'}
        ),
    ])
    def test_circular_reference_cell(self, base_workbook, set_up, values, answer):
        values = values or {}
        tbl = base_workbook['Parameters and inner links']['sht2_tbl1']
        tbl.set_records(set_up, field='fml')
        tbl.recalculate(recalc=True)
        if values:
            tbl.set_records(values, field='fml')
            tbl.recalculate(recalc=True)
        assert (tbl.data.loc[answer.keys()].value.apply(str) == pd.Series(answer)).all()
        keys = list(set_up.keys() | values.keys())
        codes = tbl.data.loc[keys].code.tolist()
        dependents = tbl.parent.all_dependents({tbl: set(codes)}, with_links=True)
        assert all(
            val == CIRCULAR_REF
            for ftbl, codes in dependents.items()
            if not (df := ftbl.data).empty
            for val in df.loc[df.code.isin(codes)].value
        )

    @pytest.mark.parametrize("set_up, values, answer", [
        (dict(F13='=+F15+F12'), dict(F12=1000),  # Referencia explícita a celda vacía.
        {'F12': '1000', 'F13': '1015', 'G16': '3428', 'G17': '3491', 'H16': '3508', 'H17': '3596'}
        ),
        (dict(F13='=+F15+SUM(F12:H12)'), dict(F12=1000),  # Referencia a celda vacía en extremo de rango.
        {'F12': '1000', 'F13': '1015', 'G16': '3428', 'G17': '3491', 'H16': '3508', 'H17': '3596'}
        ),
        (dict(F13='=+F15+SUM(F12:H12)'), dict(H12=1000),  # Referencia a celda vacía en extremo de rango.
        {'H12': '1000', 'F13': '1015', 'G16': '3428', 'G17': '3491', 'H16': '3508', 'H17': '3596'}
        ),
        (dict(F13='=+F15+SUM(F12:H12)'), dict(G12=1000),  # Referencia a celda vacía en extremo de rango.
        {'G12': '1000', 'F13': '1015', 'G16': '3428', 'G17': '3491', 'H16': '3508', 'H17': '3596'}
        ),
    ])
    def test_empty_cell_reference(self, base_workbook, set_up, values, answer):
        tbl = base_workbook['Parameters and inner links']['sht2_tbl2']
        tbl.set_records(set_up, field='fml')
        tbl.recalculate(recalc=True)
        key = list(values.keys())[0]
        # Si la referencia a la celda vacía se halla explícitamente en la fórmula o como extremo
        # en un rango de celdas, se crea el registro (record) correspondiente por lo que key in index,
        # en caso contrario no se crea el registro (key not in index).
        assert key not in tbl.data.index or key in list(set_up.values())[0] 
        tbl.set_records(values, field='value')
        tbl.recalculate(recalc=True)
        assert (tbl.data.loc[answer.keys()].value.apply(str) == pd.Series(answer)).all()


    @pytest.mark.parametrize("values, answer", [
        (
           dict(F13='=+F15+F12', G12='=+E12&F12'),   # Fórmula con referencia a celda vacía.
            {'F13': '15', 'G12': 'tabla 2:', 'G16': '2428', 'G17': '2491', 'H16': '2508', 'H17': '2596'}
        ),
        (
            dict(F13='=F5 + G13', F15='=+F14+G14'),   # Creación de nuevo enlace externo
            {'F13': '1300', 'F15': '35', 'F17': '125', 'G16': '3713', 'G17': '3776', 'H15': '73', 'H16': '3793', 'H17': '3901'}
        ),
        (
            dict(H13='=+$G$2*F13 + 75', F17='=+F16+F15+2*F14', G16='=F9 + F13 + F14'),  # Cambiar fórmulas en celdas existentes  
            {'F17': '115', 'G16': '2448', 'G17': '2511', 'H13': '75', 'H16': '2528', 'H17': '2616'}
        ),
        (
            dict(F12='=1+2', G12='=G14+F17', H12='=F12+G12'),  # Asignar fórmulas en celdas vacías 
            {'F12': '3', 'G12': '130', 'H12': '133'}
        ),
        (
            dict(F13='=1+2', G14='=F13+G15'),  # Asignar fórmulas en celdas con valores
            {'F13': '3', 'G14': '41', 'G16': '2416', 'G17': '2495', 'H14': '51', 'H16': '2496', 'H17': '2600'}
        ),
    ])
    def test_field_fml(self, values, answer, base_workbook):
        field = 'fml'
        tbl = base_workbook['Parameters and inner links']['sht2_tbl2']
        tbl.set_records(values, field=field)
        tbl.recalculate(recalc=True)
        assert (tbl.data.loc[answer.keys()].value.apply(str) == pd.Series(answer)).all()


    @pytest.mark.parametrize("values, answer", [
        (
            dict(F12=35, G12=80, H12=22),  # Asignar valores en celdas vacías
            {'F12': '35', 'G12': '80', 'H12': '22'}
        ),
        (
            dict(F13=55, F15=80, G15=22),  # Cambiar valores en celdas existentes
            {'F13': '55', 'F15': '80', 'F17': '170', 'G15': '22', 'G16': '2468', 'G17': '2515', 'H15': '102', 'H16': '2548', 'H17': '2685'}
        ),
        (
            dict(G16=333, H13=222),  # Asignar valores en celdas con fórmulas
            {'G16': '333', 'G17': '396', 'H13': '222', 'H16': '413', 'H17': '501'}
        )
    ])
    def test_field_value(self, values, answer, base_workbook):
        field = 'value'
        tbl = base_workbook['Parameters and inner links']['sht2_tbl2']
        tbl.set_records(values, field=field)
        tbl.recalculate(recalc=True)
        assert (tbl.data.loc[answer.keys()].value.apply(str) == pd.Series(answer)).all()

