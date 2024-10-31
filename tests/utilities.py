import ast
import operator
import pandas as pd
import openpyxl as px

import data
from excel_workbook import (
    ExcelTable, ExcelWorkbook, 
    coords_from_range,
    cell_pattern, cell_address, 
    EMPTY_CELL, 
)


class TableComparator:
    sort_key = staticmethod(lambda x: '{0}{2: >4s}{1: >3s}'.format(*cell_pattern.match(x).groups()))
    sort_fnc = staticmethod(lambda x: x.map(lambda x: '{0}{2: >4s}{1: >3s}'.format(*cell_pattern.match(x).groups())))

    def __init__(self, df):
        equiv = df.code.to_frame()
        self.df = self.decode_df(df, equiv)

    def difference(self, other, fields=None):
        df = self.__sub__(other, fields=fields)
        return df
    
    def symmetric_difference(self, other, fields=None):
        return self.__xor__(other, fields=fields)

    def __or__(self, other, fields=None):
        df = (
            pd.concat([self.df, other.df])
            .drop_duplicates(subset=fields)
            .assign(dependents=lambda db: db.dependents.apply(lambda x: eval(x) if x != 'nan' else set()))
            .sort_index(key=self.sort_fnc)
        )
        df.loc[:, 'dependents'] = df.dependents.where(~df.dependents.isna(), set())
        return df
    
    def __and__(self, other):
        x = self ^ other
        df = self.df.loc[~self.df.index.isin(x.index)]
        df.loc[:, 'dependents'] = df.dependents.where(~df.dependents.isna(), set())
        df = df.sort_index(key=self.sort_fnc)
        return df
    
    def __sub__(self, other, fields=None):
        x = self.__xor__(other, fields=fields)
        df = x.loc[x.index.isin(self.df.index)]
        df = df.drop_duplicates(subset=['code'], keep='first')
        df.loc[:, 'dependents'] = df.dependents.where(~df.dependents.isna(), set())
        df = df.sort_index(key=self.sort_fnc)
        return df
    
    def __xor__(self, other, fields=None):
        df = (
            pd.concat([self.df, other.df])
            .drop_duplicates(subset=fields, keep=False)
            .assign(dependents=lambda db: db.dependents.apply(lambda x: eval(x) if x != 'nan' else set()))
            .sort_index(key=self.sort_fnc)
        )
        df.loc[:, 'dependents'] = df.dependents.where(~df.dependents.isna(), set())
        return df
    
    def __eq__(self, other):
        df = self ^ other
        return df.empty
    
    def __ne__(self, value: object) -> bool:
        return not self.__eq__(value)
    
    @classmethod
    def rebase(cls, self_df, other_df):
        cells = list(set(self_df.index) & set(other_df.index))
        equiv = (
            pd.merge(
                self_df.loc[cells, ['code']].reset_index(), 
                other_df.loc[cells, ['code']].reset_index(), 
                on='cell'
            )
            .drop(columns=['cell'])
            .rename(columns={'code_y': 'code', 'code_x': 'cell'})
            .set_index('cell')
        )
        cells = list(set(other_df.index) - set(self_df.index))
        if cells:
            codes = other_df.loc[cells, 'code'].to_list()
            take_codes = set(self_df.code) & set(codes)
            new_codes = set(codes) - take_codes
            equiv = pd.concat(
                [
                    equiv,
                    pd.DataFrame([(code, f'{code}') for code in new_codes], columns=['code', 'cell']).set_index('cell'),
                    pd.DataFrame([(code, f'R{code}') for code in take_codes], columns=['code', 'cell']).set_index('cell')
                ]
            )
            pass
        return cls.decode_df(self_df, equiv)

    @staticmethod
    def decode_df(input_df, equiv):
        df = input_df.copy()
        if df.dependents.dtype == set:
            df.loc[:, 'dependents'] = df.dependents.apply(lambda x: str(x))

        df.loc[:, ['code', 'fml', 'dependents']] = (
            df.loc[:, ['code', 'fml', 'dependents']]
            .apply(lambda x: ExcelTable.encoder('decode', x, df=equiv))
        )
        df.loc[:, 'dependents'] = df.dependents.apply(lambda x: eval(x))
        df.loc[:, 'dependents'] = df.dependents.apply(
                    lambda items: str(sorted(items, key=lambda x: TableComparator.sort_key(x)))
                )
        return df




def excel_df(tbl_str, top_left_cell):
    '''
    Convierte el copy/paste de una tabla de excel a un pd.Series donde el index son las
    direcciones de las celdas y el valor es el str(value_cell)
    tbl_str: str. Copy/paste tabla excel.
    row_ini: str. Fila (row) superior izquierda de la tabla.
    col_ini: str. Columna (col) superior izquierda de la tabla.
    '''
    row_ini, col_ini = cell_address(top_left_cell)
    lst = [x.split('\t') for x in tbl_str.strip().split('\n')]
    ndx1, ndx2 = int(row_ini), int(row_ini) + len(lst)   
    col1, col2 = ord(col_ini), ord(col_ini) + len(lst[0])
    df = (
        pd.DataFrame(lst, index=pd.RangeIndex(ndx1, ndx2), columns=[chr(x) for x in range(col1, col2)])
        .unstack()
        .to_frame()
        .rename(columns={0:'value'})
        .assign(cell=lambda db: db.index.map(lambda x: x[0] + str(x[1])))
        .set_index('cell')
    )
    return df.value


def tbl_data(tbl_name, tbl_range):
    cell = tbl_range.split(':')[0]
    tblv, tblf = map(
        lambda x: excel_df(getattr(data, x), cell), 
        [f'{tbl_name}_vals', f'{tbl_name}_fmls']
    )
    fmls = tblf[tblf != tblv]
    vals = tblv.loc[~tblv.index.isin(fmls.index)].map(
        lambda x: eval(x) if x and set(x).issubset('-1234567890.') else (x or EMPTY_CELL)
        )
    empty_cells = set(vals[vals == EMPTY_CELL].index)
    vals = vals[vals != EMPTY_CELL]
    tblv = tblv[~tblv.index.isin(empty_cells)]
    return fmls.to_dict(), vals.to_dict(), tblv


def data_in_range(openpyxl_sheet, ws_range:str):
    ws = openpyxl_sheet
    rgn_coords = coords_from_range(ws_range)
    rgn_fmls = {}
    rgn_values = {}
    rgn_map = [rgn_values, rgn_fmls]
    [
        operator.setitem(rgn_map[k], cell.coordinate, value.replace('=+', '=') if k == 1 else value)
        for row in ws.iter_rows(**rgn_coords)
        for cell in row
        if (value:=cell.value) and (k:=int(isinstance(value, str) and value[0] == '=')) < 2
    ]
    return rgn_fmls, rgn_values


def wb_from_excelfile(filename, wb_structure):
    wb = px.load_workbook(filename)

    excel_wb = ExcelWorkbook('excel_module_test')
    
    for ws_name, tbl_name, tbl_range in wb_structure:
        if not ws_name in excel_wb.sheetnames:
            wsheet = excel_wb.create_worksheet(ws_name)
        fmls, values = data_in_range(wb[ws_name], tbl_range)
        sh1_tbl1 = ExcelTable(wsheet, tbl_name, tbl_range, fmls, values)
    wb.close()
    return excel_wb

